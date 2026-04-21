"""
renders.py — All render_* functions for the main predictor page

UI/UX improvements V2:
  - Animated pulse dot on fresh data badge in header
  - Score tier labels (Elite / Strong / Good / Moderate) on cards
  - Better section dividers and visual hierarchy
  - Confidence badge on Best Per Target cards
  - Statcast indicator badge on results header when data present
  - .map() replaces deprecated .applymap() (pandas 3.0)
  - width='stretch' replaces use_container_width (Streamlit 1.45+)
"""

import streamlit as st
import pandas as pd
import altair as alt
from config import CONFIG
from helpers import grade_pill, style_grade_cell, data_freshness_badge


# ─────────────────────────────────────────────────────────────────────────────
# STALENESS WARNING
# ─────────────────────────────────────────────────────────────────────────────

def render_staleness_warning():
    """
    Show a loud warning banner when the loaded slate is from a previous day.
    Reads st.session_state['slate_stale'] set by loader.load_matchups().
    """
    if not st.session_state.get('slate_stale', False):
        return

    slate_date = st.session_state.get('slate_date', 'unknown date')
    st.markdown(f"""
    <div style="background:#1c0000;border:2px solid #ef4444;border-radius:10px;
                padding:.85rem 1.2rem;margin:.5rem 0;display:flex;align-items:center;gap:.8rem;">
      <span style="font-size:1.4rem">⚠️</span>
      <div>
        <div style="font-weight:700;color:#f87171;font-size:.9rem">
          Stale Data Detected — Slate from {slate_date}
        </div>
        <div style="color:#fca5a5;font-size:.78rem;margin-top:.2rem">
          Today's Matchups.csv has not been uploaded yet. All scores reflect yesterday's
          slate. Do not place bets until today's data is loaded.
        </div>
      </div>
      <div style="margin-left:auto;font-size:.72rem;color:#f87171">
        Hit 🔄 Refresh when today's file is ready
      </div>
    </div>
    """, unsafe_allow_html=True)


def _score_tier(score: float) -> tuple[str, str]:
    """Return (label, css_color) for a 0-100 score."""
    if score >= 75: return "ELITE",    "#4ade80"
    if score >= 60: return "STRONG",   "#a3e635"
    if score >= 45: return "GOOD",     "#fbbf24"
    if score >= 30: return "MODERATE", "#fb923c"
    return             "WEAK",     "#f87171"


def _eligible_for_target(df: pd.DataFrame, sc: str) -> pd.DataFrame:
    """
    Hard profile filter — removes players whose score is higher in a
    more powerful category than the target category.

    Single:  exclude if XB_Score > Single_Score or HR_Score > Single_Score
    XB:      exclude if HR_Score > XB_Score
    Hit/HR:  no exclusions

    Falls back to the full df if filtering leaves it empty.
    """
    sc_base = sc.replace('_gc', '')

    if sc_base == 'Single_Score':
        mask = pd.Series(True, index=df.index)
        if 'XB_Score' in df.columns:
            mask &= ~(df['XB_Score'] > df['Single_Score'])
        if 'HR_Score' in df.columns:
            mask &= ~(df['HR_Score'] > df['Single_Score'])
        filtered = df[mask]
        return filtered if not filtered.empty else df

    if sc_base == 'XB_Score':
        mask = pd.Series(True, index=df.index)
        if 'HR_Score' in df.columns:
            mask &= ~(df['HR_Score'] > df['XB_Score'])
        filtered = df[mask]
        return filtered if not filtered.empty else df

    return df


def _profile_badge(row: pd.Series, target_sc: str) -> str:
    """
    Returns an HTML warning badge when the player's contact profile is a
    mismatch for the target bet type.

    Single: warn when XB_Score or HR_Score significantly exceeds Single_Score
    XB:     warn when HR_Score significantly exceeds XB_Score
    Hit/HR: no warning
    """
    if target_sc not in ('Single_Score', 'XB_Score'):
        return ""

    try:
        single = float(row.get('Single_Score', 0) or 0)
        xb     = float(row.get('XB_Score',     0) or 0)
        hr     = float(row.get('HR_Score',      0) or 0)
    except (ValueError, TypeError):
        return ""

    badge = ""

    if target_sc == 'Single_Score':
        xb_gap = xb - single
        hr_gap = hr - single
        if xb_gap >= 12:
            badge = (
                '<span style="background:#1c1400;color:#f59e0b;padding:1px 7px;'
                'border-radius:20px;font-size:.62rem;font-weight:700;margin-left:.3rem;'
                'font-family:\'JetBrains Mono\',monospace">⚡ XB PROFILE</span>'
            )
        elif xb_gap >= 7:
            badge = (
                '<span style="background:#1c1000;color:#fb923c;padding:1px 7px;'
                'border-radius:20px;font-size:.62rem;font-weight:700;margin-left:.3rem;'
                'font-family:\'JetBrains Mono\',monospace">⚡ XB LEAN</span>'
            )
        elif hr_gap >= 15:
            badge = (
                '<span style="background:#1c0000;color:#f87171;padding:1px 7px;'
                'border-radius:20px;font-size:.62rem;font-weight:700;margin-left:.3rem;'
                'font-family:\'JetBrains Mono\',monospace">💣 POWER PROFILE</span>'
            )

    elif target_sc == 'XB_Score':
        hr_gap = hr - xb
        if hr_gap >= 12:
            badge = (
                '<span style="background:#1c0000;color:#f87171;padding:1px 7px;'
                'border-radius:20px;font-size:.62rem;font-weight:700;margin-left:.3rem;'
                'font-family:\'JetBrains Mono\',monospace">💣 HR PROFILE</span>'
            )

    return badge


def render_source_status_panel(status_map: dict):
    if not status_map:
        return
    labels = {
        'matchups_csv': 'Matchups',
        'pitcher_context': 'Pitchers',
        'game_conditions': 'Game Cond',
        'quality_starts': 'QS',
        'batting_order': 'Lineups',
        'recent_form': 'Form',
        'pitcher_handedness': 'Handedness',
        'statcast': 'Statcast',
        'bvp_splits': 'BvP/Splits',
        'prop_odds': 'Prop Odds',
    }
    color_map = {'loaded': 'var(--hit)', 'empty': 'var(--xb)', 'missing': 'var(--hr)', 'unavailable': 'var(--hr)'}
    parts = []
    for key, label in labels.items():
        if key not in status_map:
            continue
        state = status_map[key]
        color = color_map.get(state, 'var(--muted)')
        parts.append(
            f'<span style="display:inline-flex;align-items:center;gap:.35rem;margin:.15rem .4rem .15rem 0;padding:.25rem .55rem;border:1px solid rgba(255,255,255,.08);border-radius:999px;background:rgba(255,255,255,.03);font-size:.7rem;color:var(--text)"><span style="width:.48rem;height:.48rem;border-radius:50%;background:{color};display:inline-block"></span>{label}: {state}</span>'
        )
    if parts:
        st.markdown('<div style="margin:.35rem 0 .65rem 0"><div style="font-size:.74rem;color:var(--muted);margin-bottom:.25rem">Source health</div>' + ''.join(parts) + '</div>', unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────────────────────────────────────

def render_header():
    badge = data_freshness_badge()
    # Add animated pulse to green badge
    badge_html = badge.replace(
        'sbadge-green">',
        'sbadge-green"><span class="pulse-dot"></span>'
    )
    st.markdown(f"""
    <div class="app-header">
      <div>
        <img src="https://github.com/a1faded/a1picks-hits-bot/blob/main/a1sports.png?raw=true"
             style="height:40px;width:auto;filter:drop-shadow(0 0 8px rgba(59,130,246,.4))" />
      </div>
      <div class="title-wrap">
        <h1>A1PICKS MLB Hit Predictor</h1>
        <p>BallPark Pal simulations &nbsp;·&nbsp; MLB Stats API &nbsp;·&nbsp; Statcast &nbsp;·&nbsp; V5.5</p>
      </div>
      <div class="meta">{badge_html}</div>
    </div>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# STAT BAR
# ─────────────────────────────────────────────────────────────────────────────

def render_stat_bar(df: pd.DataFrame):
    if df is None or df.empty:
        return

    avg_hp = df['total_hit_prob'].mean()
    avg_hr = df['p_hr'].mean() if 'p_hr' in df.columns else 0
    has_sc = any(c in df.columns and df[c].notna().any() for c in ['Barrel%', 'HH%', 'xBA'])

    lineup_badge = ""
    try:
        from mlb_api import get_lineup_status_map
        smap      = get_lineup_status_map()
        confirmed = sum(1 for v in smap.values() if '✅' in v['status'])
        total     = len(smap)
        color     = "var(--hit)" if confirmed == total else "var(--xb)"
        lineup_badge = (
            f'<div class="stat-item">'
            f'<span class="val" style="color:{color}">{confirmed}/{total}</span>'
            f'<span class="lbl">Lineups ✅</span></div>'
        )
    except Exception:
        pass

    sc_badge = (
        '<div class="stat-item">'
        '<span class="val" style="color:var(--accent);font-size:.85rem">⚡</span>'
        '<span class="lbl">Statcast ON</span></div>'
    ) if has_sc else ""

    # ── Game Conditions quick indicators ─────────────────────────────────────
    # Show slate-level GC environment as color-coded pills in the stat bar.
    # Average across all games so the user sees the overall day's environment.
    gc_badge = ""
    _GC_NEEDED = ['gc_hits20','gc_k20','gc_hr4']
    if all(c in df.columns for c in _GC_NEEDED):
        from config import CONFIG as _cfg
        avg_hits20 = df['gc_hits20'].mean()
        avg_k20    = df['gc_k20'].mean()
        avg_hr4    = df['gc_hr4'].mean()

        # Hits: above median = hitter day (green for overs, no color for unders)
        hits_color = "var(--hit)" if avg_hits20 > _cfg['gc_hits20_anchor'] else "var(--muted)"
        # K%: above median = pitcher day
        k_color    = "var(--xb)"  if avg_k20    > _cfg['gc_k20_anchor']    else "var(--muted)"
        # HR: above median = power day
        hr_color   = "var(--hr)"  if avg_hr4    > _cfg['gc_hr4_anchor']    else "var(--muted)"

        gc_badge = (
            f'<div class="stat-item">'
            f'<span class="val" style="font-size:.72rem;color:{hits_color}">'
            f'{avg_hits20:.0f}%</span>'
            f'<span class="lbl">Hit Env</span></div>'
            f'<div class="stat-item">'
            f'<span class="val" style="font-size:.72rem;color:{k_color}">'
            f'{avg_k20:.0f}%</span>'
            f'<span class="lbl">K Env</span></div>'
            f'<div class="stat-item">'
            f'<span class="val" style="font-size:.72rem;color:{hr_color}">'
            f'{avg_hr4:.0f}%</span>'
            f'<span class="lbl">HR Env</span></div>'
        )

    st.markdown(f"""
    <div class="stat-bar">
      <div class="stat-item">
        <span class="val">{len(df)}</span>
        <span class="lbl">Matchups</span>
      </div>
      <div class="stat-item">
        <span class="val">{df['Batter'].nunique()}</span>
        <span class="lbl">Batters</span>
      </div>
      <div class="stat-item">
        <span class="val">{df['Team'].nunique()}</span>
        <span class="lbl">Teams</span>
      </div>
      <div class="stat-item">
        <span class="val" style="color:var(--hit)">{avg_hp:.1f}%</span>
        <span class="lbl">Avg Hit Prob</span>
      </div>
      <div class="stat-item">
        <span class="val" style="color:var(--hr)">{avg_hr:.2f}%</span>
        <span class="lbl">Avg HR Prob</span>
      </div>
      {gc_badge}{lineup_badge}{sc_badge}
    </div>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# SCORE SUMMARY CARDS
# ─────────────────────────────────────────────────────────────────────────────

def render_score_summary_cards(slate_df: pd.DataFrame, filters: dict):
    if slate_df.empty:
        return

    st.markdown('<div class="section-head">🏆 Today\'s Best — Full Slate</div>',
                unsafe_allow_html=True)

    defs = [
        ('Hit_Score',    'scard-hit',    '🎯', 'HIT',      'Any Base Hit'),
        ('Single_Score', 'scard-single', '1️⃣', 'SINGLE',  'Single Specifically'),
        ('XB_Score',     'scard-xb',     '🔥', 'XB',       'Double / Triple'),
        ('HR_Score',     'scard-hr',     '💣', 'HR',        'Home Run'),
        ('HRR_Score',    'scard-hit',    '🔴', 'H+R+RBI',  'Hits+Runs+RBIs'),
    ]

    use_gc     = filters.get('use_gc', False)
    cards_html = '<div class="score-grid">'

    for sc, css, icon, short, desc in defs:
        if sc not in slate_df.columns:
            continue
        rank_sc   = (sc + '_gc') if (use_gc and sc + '_gc' in slate_df.columns) else sc
        eligible  = _eligible_for_target(slate_df, sc)
        row       = eligible.loc[eligible[rank_sc].idxmax()]
        disp_val  = float(row[rank_sc])
        base_col = sc + '_base'
        gc_str   = " ⛅" if use_gc and rank_sc != sc else ""

        park_str = ""
        if filters['use_park'] and base_col in slate_df.columns and row.get(base_col, 0) != 0:
            delta    = row[sc] - row[base_col]
            pct      = delta / row[base_col] * 100
            park_str = f" · <span style='color:{'var(--pos)' if delta>=0 else 'var(--neg)'}'>{'+' if delta>=0 else ''}{pct:.0f}% park</span>"

        tier_lbl, tier_color = _score_tier(disp_val)
        profile_badge = _profile_badge(row, sc)

        cards_html += f"""
        <div class="scard {css}">
          <div class="sc-type"><span>{icon} {short}</span>{gc_str} — {desc}</div>
          <div class="sc-name">{row['Batter']}{profile_badge}</div>
          <div class="sc-meta">{row['Team']} vs {row['Pitcher']}{park_str}</div>
          <div class="sc-score">{disp_val:.1f}</div>
          <div style="margin-top:.35rem">
            <span style="font-size:.62rem;padding:1px 6px;border-radius:20px;
              background:{tier_color}22;color:{tier_color};font-weight:700;
              font-family:'JetBrains Mono',monospace;letter-spacing:.04em">{tier_lbl}</span>
          </div>
        </div>"""

    cards_html += '</div>'
    st.markdown(cards_html, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# PITCHER LANDSCAPE
# ─────────────────────────────────────────────────────────────────────────────

def render_pitcher_landscape(pitcher_df, df: pd.DataFrame):
    def _fmt_pct(val, default):
        num = pd.to_numeric(pd.Series([val]), errors='coerce').iloc[0]
        return f"{float(num):.1f}%" if pd.notna(num) else f"{default:.1f}% *"

    def _fmt_mult(val, default=1.0):
        num = pd.to_numeric(pd.Series([val]), errors='coerce').iloc[0]
        return f"{float(num):.3f}×" if pd.notna(num) else f"{default:.3f}×"

    with st.expander("⚾ Pitcher Landscape", expanded=False):
        if pitcher_df is None or pitcher_df.empty:
            st.markdown('<div class="notice notice-info">ℹ️ Pitcher CSV data unavailable.</div>',
                        unsafe_allow_html=True)
            return

        today_pitchers = df['Pitcher'].astype(str).dropna().unique()
        rows_html      = ""

        # Duplicate last names can exist in the pitcher CSV inputs.
        # For the landscape panel, only use unambiguous pitcher rows; otherwise
        # fall back to neutral display values instead of crashing on Series/DataFrame formatting.
        if '_ambiguous' in pitcher_df.columns:
            landscape_df = pitcher_df[~pitcher_df['_ambiguous']].copy()
        else:
            vc = pitcher_df['last_name'].astype(str).value_counts()
            landscape_df = pitcher_df.loc[pitcher_df['last_name'].astype(str).map(vc).fillna(0).eq(1)].copy()
        pm = landscape_df.drop_duplicates(subset=['last_name']).set_index('last_name') if not landscape_df.empty else pd.DataFrame()

        for p in sorted(today_pitchers):
            if not pm.empty and p in pm.index:
                r       = pm.loc[p]
                grade_h = grade_pill(str(r.get('pitch_grade', 'B')))
                name, team = r.get('full_name', p), r.get('team', '—')
                hit_val    = _fmt_pct(r.get('hit8_prob'), CONFIG['pitcher_hit_neutral'])
                hr_val     = _fmt_pct(r.get('hr2_prob'), CONFIG['pitcher_hr_neutral'])
                wk_val     = _fmt_pct(r.get('walk3_prob'), CONFIG['pitcher_walk_neutral'])
                hm_val     = _fmt_mult(r.get('pitch_hit_mult'), 1.0)
                hrm_val    = _fmt_mult(r.get('pitch_hr_mult'), 1.0)
            else:
                grade_h = grade_pill('B')
                name, team = p, "—"
                hit_val = f"{CONFIG['pitcher_hit_neutral']:.1f}% *"
                hr_val  = f"{CONFIG['pitcher_hr_neutral']:.1f}% *"
                wk_val  = f"{CONFIG['pitcher_walk_neutral']:.1f}% *"
                hm_val  = hrm_val = "1.000×"

            rows_html += f"""<tr>
              <td style="color:var(--text)">{name}</td>
              <td style="color:var(--muted)">{team}</td>
              <td>{grade_h}</td>
              <td style="color:var(--hit)">{hit_val}</td>
              <td style="color:var(--hr)">{hr_val}</td>
              <td style="color:var(--xb)">{wk_val}</td>
              <td style="color:var(--muted)">{hm_val}</td>
              <td style="color:var(--muted)">{hrm_val}</td>
            </tr>"""

        st.markdown(f"""
        <div class="pt-wrap"><table class="pt-table"><thead><tr>
          <th>Pitcher</th><th>Team</th><th>Grade</th>
          <th>Hit 8+</th><th>HR 2+</th><th>Walk 3+</th>
          <th>Hit Mult</th><th>HR Mult</th>
        </tr></thead><tbody>{rows_html}</tbody></table></div>
        <div class="notice notice-pitcher" style="margin-top:.5rem">
          📊 <b>Hit 8+</b> drives Hit/Single/XB multiplier · <b>HR 2+</b> drives HR multiplier ·
          <b>Walk 3+</b> mild penalty all scores · Max effect ±8%
        </div>""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# PARK NOTICE
# ─────────────────────────────────────────────────────────────────────────────

def render_park_notice(slate_df: pd.DataFrame, filters: dict):
    sc_base  = filters.get('score_col_base', filters['score_col'].replace('_gc', ''))
    use_gc   = filters.get('use_gc', False)
    sc       = (sc_base + '_gc') if (use_gc and sc_base + '_gc' in slate_df.columns) else sc_base
    base_sc  = sc_base + '_base'

    if not filters['use_park']:
        st.markdown('<div class="notice notice-park">🏟️ <b>Park OFF</b> — pure player vs pitcher.</div>',
                    unsafe_allow_html=True)
        return

    if base_sc not in slate_df.columns or slate_df.empty:
        return

    ab, bl = slate_df[base_sc].mean(), slate_df[sc].mean()
    delta   = bl - ab
    pct     = (delta / ab * 100) if ab != 0 else 0
    dir_    = "boosted" if delta >= 0 else "reduced"
    col_    = "var(--pos)" if delta >= 0 else "var(--neg)"

    st.markdown(
        f'<div class="notice notice-park">🏟️ <b>Park ON</b> — '
        f'<span style="color:{col_};font-weight:700">{dir_} ~{abs(pct):.1f}%</span> avg · '
        f'Park Δ column shows per-player impact</div>',
        unsafe_allow_html=True
    )


# ─────────────────────────────────────────────────────────────────────────────
# GAME CONDITIONS PANEL
# ─────────────────────────────────────────────────────────────────────────────

def render_game_conditions_panel(slate_df: pd.DataFrame, filters: dict,
                                  game_cond, pitcher_qs):
    use_gc  = filters.get('use_gc', False)
    sc_base = filters.get('score_col_base', filters['score_col'].replace('_gc', ''))
    sc      = (sc_base + '_gc') if (use_gc and sc_base + '_gc' in slate_df.columns) else sc_base

    if not use_gc:
        return

    gc_cols = ['gc_hr4', 'gc_hits20', 'gc_k20', 'gc_walks8', 'gc_runs10', 'gc_qs']
    if slate_df.empty or not all(c in slate_df.columns for c in gc_cols):
        st.markdown(
            '<div class="notice notice-warn">🌦️ <b>Game Conditions ON</b> — '
            'No game condition CSVs found. Scores unaffected.</div>',
            unsafe_allow_html=True)
        return

    gc_sc_col = sc if sc.endswith("_gc") else sc + "_gc"
    if gc_sc_col not in slate_df.columns:
        return

    game_rows = []
    for game in sorted(slate_df['Game'].unique()):
        gdf = slate_df[slate_df['Game'] == game]
        if gdf.empty:
            continue
        row      = gdf.iloc[0]
        avg_base = gdf[sc].mean()
        avg_gc   = gdf[gc_sc_col].mean()
        game_rows.append({
            'Game':         game,
            '4+ HR %':     f"{row['gc_hr4']:.1f}%",
            '20+ Hits %':  f"{row['gc_hits20']:.1f}%",
            '20+ Ks %':    f"{row['gc_k20']:.1f}%",
            '8+ Walks %':  f"{row['gc_walks8']:.1f}%",
            '10+ Runs %':  f"{row['gc_runs10']:.1f}%",
            'QS %':         f"{row['gc_qs']:.1f}%",
            'Cond Δ (avg)': avg_gc - avg_base,
        })

    if not game_rows:
        return

    sc_lbl = {'Hit_Score':'Hit','Single_Score':'Single',
               'XB_Score':'XB','HR_Score':'HR',
               'HRR_Score':'H+R+RBI'}.get(sc, 'Score')

    with st.expander(f"🌦️ Game Conditions — {sc_lbl} Score Impact", expanded=True):
        gdf_disp = pd.DataFrame(game_rows)
        styled   = gdf_disp.style.format({'Cond Δ (avg)': '{:+.1f}'})
        styled   = styled.background_gradient(subset=['Cond Δ (avg)'],
                                              cmap='RdYlGn', vmin=-8, vmax=8)
        st.dataframe(styled, width='stretch', hide_index=True)


# ─────────────────────────────────────────────────────────────────────────────
# RESULTS TABLE
# ─────────────────────────────────────────────────────────────────────────────

def _build_export_df(filtered_df: pd.DataFrame, filters: dict) -> pd.DataFrame:
    """
    Build a clean, human-readable DataFrame for CSV export.
    Uses the same column selection and renaming logic as render_results_table
    but returns a formatted df instead of a styled one.

    - Internal columns (_*) excluded
    - Raw multiplier columns excluded
    - Numeric columns rounded to appropriate precision
    - Column names match what users see in the table
    - Tier column included for offline reference
    """
    if filtered_df.empty:
        return pd.DataFrame()

    use_gc   = filters.get('use_gc', False)
    use_park = filters.get('use_park', True)
    sc_base  = filters.get('score_col_base', filters['score_col'].replace('_gc', ''))
    sc_gc    = sc_base + '_gc'
    base_sc  = sc_base + '_base'
    sc       = sc_gc if (use_gc and sc_gc in filtered_df.columns) else sc_base

    df = filtered_df.copy()

    # Remove internal/raw columns
    drop_pats = ['_order_pos','_form_rate','_form_label','_pitcher_hand',
                 'pitch_hit_mult','pitch_hr_mult','pitch_walk_pen',
                 'rc_norm','rc_contrib','vs_mod','vs_contrib',
                 'xb_boost','hist_bonus','p_1b_base','p_xb_base',
                 'p_hr_base','p_k_base','p_bb_base']
    df = df.drop(columns=[c for c in drop_pats if c in df.columns], errors='ignore')
    df = df[[c for c in df.columns if not c.startswith('_') and
             not c.startswith('disq') and c not in
             ('Starter','p_1b_park','p_xb_park','p_hr_park')]]

    # Compute display columns
    if 'p_k' in df.columns:
        df['K% ↓Lg']  = (CONFIG['league_k_avg']  - df['p_k']).round(1)
    if 'p_bb' in df.columns:
        df['BB% ↓Lg'] = (CONFIG['league_bb_avg'] - df['p_bb']).round(1)
    if 'p_hr' in df.columns:
        df['HR% ↑Lg'] = (df['p_hr'] - CONFIG['league_hr_avg']).round(2)
    if 'total_hit_prob' in df.columns:
        df['Hit%']    = df['total_hit_prob'].round(1)
    if 'PA' in df.columns:
        df['PA'] = pd.to_numeric(df['PA'], errors='coerce').fillna(0).astype(int)
    if 'AVG' in df.columns:
        df['AVG'] = pd.to_numeric(df['AVG'], errors='coerce').round(3)

    # Score tier
    lbl = {'Hit_Score':'Hit Score','Single_Score':'Single Score',
           'XB_Score':'XB Score','HR_Score':'HR Score','HRR_Score':'H+R+RBI Score'}
    active_label = lbl.get(sc_base, 'Score')
    if sc in df.columns:
        tier_map = lambda s: (
            "ELITE"    if s >= 75 else "STRONG" if s >= 60 else
            "GOOD"     if s >= 45 else "MODERATE" if s >= 30 else "WEAK"
        )
        tier_vals = df[sc].apply(lambda x: tier_map(float(x)) if pd.notna(x) else "")

    # Column rename map — mirrors render_results_table naming
    rename = {
        sc:            active_label,
        sc_base:       f"{active_label} (Base)",
        sc_gc:         f"{active_label} (GC)",
        'pitch_grade': 'Pitcher Grade',
        'p_k':         'K%', 'p_bb':'BB%',
        'p_1b':        '1B%','p_xb':'XB%','p_hr':'HR%',
        'total_hit_prob':'Hit%',
        'vs Grade':    'vs Pitcher Grade',
        'Hit_Score':   'Hit Score', 'Single_Score':'Single Score',
        'XB_Score':    'XB Score',  'HR_Score':'HR Score',
        'HRR_Score':   'H+R+RBI Score',
        'Hit_Score_gc':'Hit Score (GC)', 'XB_Score_gc':'XB Score (GC)',
        'HR_Score_gc': 'HR Score (GC)',  'HRR_Score_gc':'H+R+RBI Score (GC)',
        'Hit_Score_base':'Hit Score (Base)', 'XB_Score_base':'XB Score (Base)',
        'bvp_avg':     'BvP Career AVG',  'bvp_ops':'BvP Career OPS',
        'bvp_ab':      'BvP Career AB',   'bvp_hr':'BvP Career HR',
        'bvp_rbi':     'BvP Career RBI',  'bvp_k':'BvP Career K',
        'split_avg':   'Split AVG (vs hand)', 'split_ops':'Split OPS (vs hand)',
        'prop_tb_line':'TB Line', 'prop_tb_under_odds':'TB Under Odds',
        'prop_tb_over_odds':'TB Over Odds', 'prop_hr_odds':'HR Odds',
    }
    df = df.rename(columns={k:v for k,v in rename.items() if k in df.columns})

    # Insert Tier column right after active score column
    if active_label in df.columns and tier_vals is not None:
        insert_at = df.columns.get_loc(active_label) + 1
        df = df.copy()
        df.insert(insert_at, 'Tier', tier_vals)

    # Round all remaining float columns to sensible precision
    for col in df.columns:
        if col in ('Tier','Pitcher Grade','Batter','Team','Pitcher',
                   'TB Line','TB Under Odds','TB Over Odds','HR Odds'):
            continue
        try:
            series = pd.to_numeric(df[col], errors='coerce')
            if series.notna().any():
                if 'AVG' in col or 'OPS' in col or col in ('BvP Career AVG','Split AVG (vs hand)'):
                    df[col] = series.round(3)
                elif col.endswith('%') or 'Score' in col or col in ('Tier',):
                    df[col] = series.round(1)
                else:
                    df[col] = series.round(2)
        except Exception:
            pass

    return df


def _build_export_xlsx(filtered_df: pd.DataFrame, filters: dict) -> bytes:
    """
    Build a formatted Excel (.xlsx) workbook from the current results.

    Formatting applied:
    - Frozen header row, auto-filter on all columns
    - Bold, dark-background header row with white text
    - Tier column color-coded (ELITE=green, STRONG=yellow, GOOD=orange, etc.)
    - Score columns with conditional color scale (white→green gradient)
    - BvP AVG/OPS columns color-coded (green=dominant, red=struggles)
    - Alternating row shading for readability
    - Auto-fitted column widths
    - Separate 'Under Scores' sheet when under score columns present
    """
    import io
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

    # Build clean df first
    export_df = _build_export_df(filtered_df, filters)
    if export_df.empty:
        return b""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "A1PICKS Results"

    # ── Color palette ─────────────────────────────────────────────────────────
    C_HEADER_BG  = "0D1321"   # dark navy header
    C_HEADER_FG  = "E8EEF7"   # light text
    C_ALT_ROW    = "0F1923"   # slightly lighter than header
    C_ELITE      = "052E16"   # dark green bg
    C_ELITE_FG   = "4ADE80"
    C_STRONG     = "1C2A00"
    C_STRONG_FG  = "A3E635"
    C_GOOD       = "1C1500"
    C_GOOD_FG    = "FBBF24"
    C_MOD        = "1C0800"
    C_MOD_FG     = "FB923C"
    C_WEAK       = "1C0000"
    C_WEAK_FG    = "F87171"
    C_BVP_HIGH   = "052E16"   # strong BvP history = green
    C_BVP_HIGH_FG= "4ADE80"
    C_BVP_LOW    = "1C0000"
    C_BVP_LOW_FG = "F87171"

    # ── Write header row ──────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor=C_HEADER_BG)
    header_font = Font(bold=True, color=C_HEADER_FG, name="Calibri", size=9)
    header_align= Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="1E2D3D"),
        right=Side(style="thin", color="1E2D3D"),
    )

    for col_idx, col_name in enumerate(export_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # ── Write data rows ───────────────────────────────────────────────────────
    alt_fill = PatternFill("solid", fgColor=C_ALT_ROW)
    base_font = Font(name="Calibri", size=9, color=C_HEADER_FG)

    for row_idx, (_, row) in enumerate(export_df.iterrows(), start=2):
        row_fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, (col_name, value) in enumerate(row.items(), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = base_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                right=Side(style="thin", color="1A2540"),
                bottom=Side(style="thin", color="1A2540"),
            )
            if row_fill:
                cell.fill = row_fill

            # Tier column — color-coded by value
            if col_name == "Tier" and isinstance(value, str):
                tier_map = {
                    "ELITE":    (C_ELITE,  C_ELITE_FG),
                    "STRONG":   (C_STRONG, C_STRONG_FG),
                    "GOOD":     (C_GOOD,   C_GOOD_FG),
                    "MODERATE": (C_MOD,    C_MOD_FG),
                    "WEAK":     (C_WEAK,   C_WEAK_FG),
                }
                if value in tier_map:
                    bg, fg = tier_map[value]
                    cell.fill = PatternFill("solid", fgColor=bg)
                    cell.font = Font(name="Calibri", size=9, color=fg, bold=True)
                    cell.alignment = Alignment(horizontal="center")

            # Score columns — number format
            elif ("Score" in col_name or "%" in col_name) and isinstance(value, (int, float)):
                cell.number_format = "0.0"

            # BvP columns — color high (>.300) green, low (<.200) red
            elif "BvP" in col_name and "AVG" in col_name and isinstance(value, (int, float)):
                cell.number_format = "0.000"
                if value >= 0.300:
                    cell.fill = PatternFill("solid", fgColor=C_BVP_HIGH)
                    cell.font = Font(name="Calibri", size=9, color=C_BVP_HIGH_FG)
                elif value > 0 and value < 0.200:
                    cell.fill = PatternFill("solid", fgColor=C_BVP_LOW)
                    cell.font = Font(name="Calibri", size=9, color=C_BVP_LOW_FG)

            elif "AVG" in col_name and isinstance(value, (int, float)):
                cell.number_format = "0.000"

            elif "OPS" in col_name and isinstance(value, (int, float)):
                cell.number_format = "0.000"

    # ── Auto-filter on header row ─────────────────────────────────────────────
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(export_df.columns))}1"
    )

    # ── Conditional color scale on main Score column ──────────────────────────
    score_col_name = None
    for col_name in export_df.columns:
        if "Score" in col_name and "BvP" not in col_name and "Base" not in col_name:
            score_col_name = col_name
            break

    if score_col_name and score_col_name in export_df.columns:
        col_letter = get_column_letter(
            list(export_df.columns).index(score_col_name) + 1)
        last_row   = len(export_df) + 1
        score_range = f"{col_letter}2:{col_letter}{last_row}"
        ws.conditional_formatting.add(
            score_range,
            ColorScaleRule(
                start_type="num",  start_value=0,  start_color="1C0000",
                mid_type="num",    mid_value=50,   mid_color="1C1500",
                end_type="num",    end_value=100,  end_color="052E16",
            )
        )

    # ── Column widths ─────────────────────────────────────────────────────────
    width_map = {
        "Batter": 16, "Team": 7, "Pitcher": 16, "Pitcher Grade": 10,
        "Tier": 9, "Profile": 12, "Pos": 5, "Form": 8,
    }
    for col_idx, col_name in enumerate(export_df.columns, start=1):
        letter = get_column_letter(col_idx)
        if col_name in width_map:
            ws.column_dimensions[letter].width = width_map[col_name]
        elif "Score" in col_name:
            ws.column_dimensions[letter].width = 13
        elif "%" in col_name or "Δ" in col_name:
            ws.column_dimensions[letter].width = 8
        elif "AVG" in col_name or "OPS" in col_name:
            ws.column_dimensions[letter].width = 11
        elif "BvP" in col_name:
            ws.column_dimensions[letter].width = 10
        elif "Odds" in col_name or "Line" in col_name:
            ws.column_dimensions[letter].width = 10
        else:
            ws.column_dimensions[letter].width = 9

    # ── Tab styling ───────────────────────────────────────────────────────────
    ws.sheet_properties.tabColor = "10B981"   # green accent

    # ── Save to bytes ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def render_results_table(filtered_df: pd.DataFrame, filters: dict):
    if filtered_df.empty:
        st.warning("⚠️ No players match your filters — try relaxing the thresholds.")
        return

    use_gc   = filters.get('use_gc', False)
    use_park = filters['use_park']
    sc_base  = filters.get('score_col_base', filters['score_col'].replace('_gc', ''))
    sc_gc    = sc_base + '_gc'
    base_sc  = sc_base + '_base'
    sc       = sc_gc if (use_gc and sc_gc in filtered_df.columns) else sc_base

    disp = filtered_df.copy()
    disp['K% ↓Lg']  = (CONFIG['league_k_avg']  - disp['p_k']).round(1)
    disp['BB% ↓Lg'] = (CONFIG['league_bb_avg'] - disp['p_bb']).round(1)
    disp['HR% ↑Lg'] = (disp['p_hr'] - CONFIG['league_hr_avg']).round(2)
    disp['Hit%']    = disp['total_hit_prob'].round(1)
    disp['PA']      = disp['PA'].astype(int)
    disp['AVG']     = disp['AVG'].round(3)
    disp['vs Grade']= pd.to_numeric(disp['vs Grade'], errors='coerce').round(0).astype(int)
    disp['Park Δ']  = (
        (disp[sc_base] - disp[base_sc]).round(1)
        if (use_park and base_sc in disp.columns) else 0.0
    )
    disp['Cond Δ'] = (
        (disp[sc_gc] - disp[sc_base]).round(1)
        if (use_gc and sc_gc in disp.columns) else 0.0
    )

    # ── New signal columns ─────────────────────────────────────────────────────
    has_order = '_order_pos' in disp.columns and disp['_order_pos'].notna().any()
    has_form  = '_form_label' in disp.columns

    if has_order:
        disp['Pos'] = disp['_order_pos'].apply(
            lambda x: f"#{int(x)}" if pd.notna(x) else "—"
        )
    if has_form:
        disp['Form'] = disp['_form_label'].fillna('—')

    lbl = {
        'Hit_Score':   '🎯 Hit',
        'Single_Score':'1️⃣ Single',
        'XB_Score':    '🔥 XB',
        'HR_Score':    '💣 HR',
        'HRR_Score':   '🔴 H+R+RBI',
    }
    active       = lbl.get(sc_base, 'Score')
    active_label = (active + ' ⛅') if (use_gc and sc_gc in filtered_df.columns) else active

    cols = {'Batter':'Batter','Team':'Team','Pitcher':'Pitcher',
            'pitch_grade':'P.Grd', sc: active_label}

    # Profile mismatch column — immediately after score for visibility
    if 'Profile' in disp.columns:
        cols['Profile'] = 'Profile'

    # Inject Pos immediately after Batter/Team when available
    if has_order and 'Pos' in disp.columns:
        cols['Pos'] = 'Pos'

    if use_park and base_sc in disp.columns:
        cols[base_sc]  = 'Base'
        cols['Park Δ'] = 'Park Δ'
    if use_gc and sc_gc in disp.columns:
        cols['Cond Δ'] = 'Cond Δ'

    for sc2, lb2 in lbl.items():
        if sc2 != sc_base and sc2 != sc_gc and sc2 in disp.columns:
            cols[sc2] = lb2

    cols.update({'Hit%':'Hit%','p_1b':'1B%','p_xb':'XB%','p_hr':'HR%',
                 'p_k':'K%','p_bb':'BB%','K% ↓Lg':'K% ↓Lg',
                 'BB% ↓Lg':'BB% ↓Lg','HR% ↑Lg':'HR% ↑Lg',
                 'vs Grade':'vsPit','PA':'PA','AVG':'AVG'})

    # Form near the end (before Statcast)
    if has_form and 'Form' in disp.columns:
        cols['Form'] = 'Form'

    # ── Profile mismatch column ────────────────────────────────────────────────
    sc_base_for_profile = filters.get('score_col_base',
                                      filters['score_col'].replace('_gc',''))
    if sc_base_for_profile in ('Single_Score', 'XB_Score'):
        def _profile_label(row):
            try:
                xb  = float(row.get('XB_Score',  0) or 0)
                hr  = float(row.get('HR_Score',   0) or 0)
                si  = float(row.get('Single_Score',0) or 0)
                if sc_base_for_profile == 'Single_Score':
                    if xb - si >= 12:  return "⚡ XB Profile"
                    if xb - si >=  7:  return "⚡ XB Lean"
                    if hr - si >= 15:  return "💣 Power"
                    return "✅ Clean"
                else:
                    if hr - xb >= 12:  return "💣 HR Profile"
                    return "✅ Clean"
            except Exception:
                return ""
        disp['Profile'] = disp.apply(_profile_label, axis=1)

    # ── Prop odds columns ──────────────────────────────────────────────────────
    has_props = 'prop_tb_line' in disp.columns and disp['prop_tb_line'].astype(bool).any()
    if has_props:
        cols['prop_tb_line']       = 'TB Line'
        cols['prop_tb_under_odds'] = 'TB Under'
        cols['prop_tb_over_odds']  = 'TB Over'
        if sc_base == 'HR_Score':
            cols['prop_hr_odds'] = 'HR Odds'

    # ── BvP columns ───────────────────────────────────────────────────────────
    # Show career stats vs today's specific pitcher when available.
    # bvp_conf tells users how reliable the sample is (0–1, shown as AB count).
    has_bvp = 'bvp_ab' in disp.columns and disp['bvp_ab'].notna().any()
    if has_bvp:
        cols['bvp_ab']  = 'BvP AB'
        cols['bvp_avg'] = 'BvP AVG'
        cols['bvp_ops'] = 'BvP OPS'
        cols['bvp_hr']  = 'BvP HR'

    # ── Splits column ─────────────────────────────────────────────────────────
    # Batter's season AVG vs this pitcher's hand (L/R)
    has_splits = 'split_avg' in disp.columns and disp['split_avg'].notna().any()
    if has_splits:
        cols['split_avg'] = 'Split AVG'
        cols['split_ops'] = 'Split OPS'

    statcast_cols = {'Barrel%':'Barrel%','HH%':'HH%','xBA':'xBA',
                     'xSLG':'xSLG','AvgEV':'AvgEV','maxEV':'maxEV'}
    has_statcast  = any(c in disp.columns and disp[c].notna().any() for c in statcast_cols)
    if has_statcast:
        for raw, label in statcast_cols.items():
            if raw in disp.columns:
                cols[raw] = label

    existing = [c for c in cols if c in disp.columns]
    out_df   = disp[existing].rename(columns=cols)

    # ── Add Tier column (ELITE/STRONG/GOOD/MODERATE/WEAK) ─────────────────────
    # Placed immediately after the score column so users get a text label
    # alongside the number — no mental calibration needed.
    tier_map = lambda s: (
        "🟢 ELITE"    if s >= 75 else
        "🟡 STRONG"   if s >= 60 else
        "🟠 GOOD"     if s >= 45 else
        "🔴 MODERATE" if s >= 30 else
        "⚫ WEAK"
    )
    if active_label in out_df.columns:
        tier_vals = out_df[active_label].apply(
            lambda x: tier_map(float(x)) if pd.notna(x) else "—"
        )
        insert_at = out_df.columns.get_loc(active_label) + 1
        out_df    = out_df.copy()       # decouple before mutation
        out_df.insert(insert_at, 'Tier', tier_vals)

    _text_cols = {'Tier', 'Profile', 'Form', 'Pos', 'Market Edge',
                  'TB Line', 'TB Under', 'TB Over', 'HR Odds', 'P.Grd',
                  'Batter', 'Team', 'Pitcher'}

    fmt = {}
    for cn in out_df.columns:
        if cn in _text_cols:
            continue                    # skip — no format for string columns
        if cn in ['Hit%','1B%','XB%','HR%','K%','BB%','Barrel%','HH%']:
            fmt[cn] = "{:.1f}%"
        elif cn in ['K% ↓Lg','BB% ↓Lg']:
            fmt[cn] = "{:+.1f}%"
        elif cn == 'HR% ↑Lg':
            fmt[cn] = "{:+.2f}%"
        elif cn in ['Park Δ','Cond Δ']:
            fmt[cn] = "{:+.1f}"
        elif cn in ['AVG','xBA','xSLG','BvP AVG','Split AVG']:
            fmt[cn] = "{:.3f}"
        elif cn in ['BvP OPS','Split OPS']:
            fmt[cn] = "{:.3f}"
        elif cn == 'BvP AB':
            fmt[cn] = "{:.0f}"
        elif cn == 'BvP HR':
            fmt[cn] = "{:.0f}"
        elif cn in ['AvgEV','maxEV']:
            fmt[cn] = "{:.1f}"
        elif any(e in cn for e in ['🎯','1️⃣','🔥','💣','Base']) and 'Prob' not in cn:
            fmt[cn] = "{:.1f}"

    styled = out_df.style.format(fmt, na_rep="—")

    for sn, cm in {'🎯 Hit':'Greens','1️⃣ Single':'GnBu',
                   '🔥 XB':'YlOrBr','💣 HR':'YlOrRd',
                   '🔴 H+R+RBI':'RdPu'}.items():
        tc = sn + ' ⛅' if (sn + ' ⛅') in out_df.columns else sn
        if tc in out_df.columns:
            try:
                styled = styled.background_gradient(subset=[tc], cmap=cm, vmin=0, vmax=100)
            except Exception:
                pass

    for cn, cm, v0, v1 in [
        ('Park Δ','RdYlGn',-10,10), ('Cond Δ','RdYlGn',-8,8),
        ('K% ↓Lg','RdYlGn',-8,12), ('HR% ↑Lg','RdYlGn',-2,3),
        ('vsPit','RdYlGn',-10,10),
        ('Barrel%','Greens',0,20),  ('HH%','Greens',20,60),
        ('AvgEV','Greens',85,100),
        # BvP — green = good history vs this pitcher
        ('BvP AVG', 'RdYlGn', 0.150, 0.400),
        ('BvP OPS', 'RdYlGn', 0.400, 1.200),
        # Splits — season avg vs pitcher hand
        ('Split AVG','RdYlGn', 0.150, 0.380),
        ('Split OPS','RdYlGn', 0.500, 1.100),
    ]:
        if cn in out_df.columns:
            try:
                styled = styled.background_gradient(subset=[cn], cmap=cm, vmin=v0, vmax=v1)
            except Exception:
                pass

    if 'P.Grd' in out_df.columns:
        styled = styled.map(style_grade_cell, subset=['P.Grd'])

    # ── Column config — widths, tooltips, number formatting ───────────────────
    col_cfg: dict = {}
    try:
        import streamlit as _st
        CC = _st.column_config

        col_cfg['Batter'] = CC.TextColumn("Batter", width="medium")
        col_cfg['Team']   = CC.TextColumn("Team",   width="small")
        col_cfg['Tier']   = CC.TextColumn(
            "Tier", width="small",
            help="ELITE ≥75 · STRONG ≥60 · GOOD ≥45 · MODERATE ≥30 · WEAK <30"
        )
        if 'Profile' in out_df.columns:
            col_cfg['Profile'] = CC.TextColumn(
                "Profile", width="small",
                help="Contact profile fit for this prop. ✅ Clean = ideal. "
                     "⚡ XB/💣 Power = player's contact tends toward extra bases."
            )
        for score_lbl in ['🎯 Hit','1️⃣ Single','🔥 XB','💣 HR',
                           '🎯 Hit ⛅','1️⃣ Single ⛅','🔥 XB ⛅','💣 HR ⛅']:
            if score_lbl in out_df.columns:
                col_cfg[score_lbl] = CC.NumberColumn(
                    score_lbl, format="%.1f",
                    min_value=0, max_value=100,
                    help="Score 0–100. Higher = stronger candidate for this prop."
                )
        for odds_col in ['TB Under','TB Over','HR Odds']:
            if odds_col in out_df.columns:
                col_cfg[odds_col] = CC.TextColumn(
                    odds_col, width="small",
                    help="American odds from Tank01 market data. e.g. -190 means you bet $190 to win $100."
                )
        if 'TB Line' in out_df.columns:
            col_cfg['TB Line'] = CC.TextColumn(
                "TB Line", width="small",
                help="Total Bases line set by the sportsbook (0.5 or 1.5)."
            )
        if 'Market Edge' in out_df.columns:
            col_cfg['Market Edge'] = CC.TextColumn(
                "Market Edge", width="small",
                help="⚡ EDGE = model favours under more than market. "
                     "✅ CONFIRMED = both agree. 🔄 CONTRARIAN = market more bullish on under than model."
            )
        # BvP columns
        for bvp_col, bvp_fmt, bvp_help in [
            ('BvP AB',  "%d",    "Career at-bats vs today's specific pitcher. ≥15=high confidence · 5-14=partial"),
            ('BvP AVG', "%.3f",  "Career batting average vs today's starting pitcher. >.280 = historically owns this pitcher"),
            ('BvP OPS', "%.3f",  "Career OPS vs today's starting pitcher. Best H+R+RBI predictor"),
            ('BvP HR',  "%d",    "Career home runs vs today's starting pitcher"),
        ]:
            if bvp_col in out_df.columns:
                col_cfg[bvp_col] = CC.NumberColumn(bvp_col, format=bvp_fmt, help=bvp_help)
        # Splits columns
        for sp_col, sp_help in [
            ('Split AVG', "Batter's season AVG vs this pitcher's hand (L/R). Current-season performance."),
            ('Split OPS', "Batter's season OPS vs this pitcher's hand."),
        ]:
            if sp_col in out_df.columns:
                col_cfg[sp_col] = CC.NumberColumn(sp_col, format="%.3f", help=sp_help)
    except Exception:
        col_cfg = {}

    st.dataframe(styled, width="stretch", column_config=col_cfg or None,
                 hide_index=False)

    LG        = CONFIG
    park_note = (
        "<b>Park Δ</b> = park impact per player."
        if use_park else "Park OFF — pure player vs pitcher."
    )
    gc_note = " · <b>Cond Δ ⛅</b> = game conditions shift" if use_gc else ""
    sc_note = " · <b>Barrel%/HH%/AvgEV</b> = Statcast season" if has_statcast else ""

    st.markdown(f"""
    <div class="legend-compact">
      <span class="hit-c">🎯 Hit</span> 1B×3 · K pen heavy &nbsp;
      <span class="sl-c">1️⃣ Single</span> 1B×5 · XB/HR penalised &nbsp;
      <span class="xb-c">🔥 XB</span> XB×5 · mod K &nbsp;
      <span class="hr-c">💣 HR</span> HR×6 · K near-neutral · BB neutral &nbsp;
      <span style="color:var(--accent)">🔴 H+R+RBI</span> Hit×0.40 + Order + HR×0.25 + GC + BvP<br>
      <b>K% ↓Lg</b> vs league {LG['league_k_avg']}% ·
      <b>BB% ↓Lg</b> vs league {LG['league_bb_avg']}% ·
      <b>HR% ↑Lg</b> vs league {LG['league_hr_avg']}% ·
      <b>PA/AVG</b> vs this pitcher ·
      {park_note}{gc_note}{sc_note}
    </div>
    """, unsafe_allow_html=True)

    # ── HRR Game Log Panel — lazy loaded for top candidates only ─────────────
    # Only shown when H+R+RBI is the active target. Fetches last 10 games for
    # the top 15 candidates AFTER scoring/filtering — never blocks page load.
    if sc_base == 'HRR_Score' and not filtered_df.empty:
        _render_hrr_game_log_panel(filtered_df, filters)


# ─────────────────────────────────────────────────────────────────────────────
# HRR GAME LOG PANEL
# ─────────────────────────────────────────────────────────────────────────────

def _render_hrr_game_log_panel(filtered_df: pd.DataFrame, filters: dict):
    """
    Lazy-loaded display panel for top HRR candidates showing recent game log
    history. Only rendered when H+R+RBI is the active target.

    Fetches last-10-games data for up to 15 top candidates via
    get_hrr_game_log_map() — cached 30 min, ~15 API calls max.
    Displayed as a compact card grid under the main table.
    """
    import streamlit as st

    sc_col = filters.get('score_col_base', 'HRR_Score')
    sc_gc  = sc_col + '_gc'
    rank_col = sc_gc if (filters.get('use_gc') and sc_gc in filtered_df.columns) \
               else sc_col

    if rank_col not in filtered_df.columns:
        return

    top_df = filtered_df.nlargest(15, rank_col)
    if top_df.empty:
        return

    st.markdown("---")
    st.markdown("#### 🔴 H+R+RBI — Recent Game Log History (Last 10 Games)")
    st.markdown(
        '<div style="font-size:.74rem;color:#64748b;margin-bottom:.5rem">'
        'How often each player accumulated ≥2 combined Hits+Runs+RBIs in their last 10 games. '
        'Loaded after scoring — shows recent form context for top candidates.</div>',
        unsafe_allow_html=True
    )

    # Resolve player IDs for top candidates
    try:
        from mlb_api import build_player_id_map
        player_id_map = build_player_id_map(
            tuple(sorted(top_df['Batter'].unique().tolist()))
        )
    except Exception:
        player_id_map = {}

    # Reverse map: id → name
    id_to_name = {v: k for k, v in player_id_map.items()}

    # Get player IDs for top candidates
    top_ids = tuple(
        player_id_map[b] for b in top_df['Batter']
        if b in player_id_map
    )

    if not top_ids:
        st.info("⏳ Player IDs not resolved — game log unavailable for this slate.")
        return

    with st.spinner("Loading recent game logs for top H+R+RBI candidates…"):
        try:
            from mlb_api import get_hrr_game_log_map
            log_map = get_hrr_game_log_map(top_ids, last_n=10)
        except Exception:
            log_map = {}

    if not log_map:
        st.info("Game log data unavailable — try again closer to game time.")
        return

    # Render as a card grid
    cards_html = '<div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(180px,1fr));gap:.5rem;margin:.4rem 0">'

    for _, row in top_df.iterrows():
        batter    = row.get('Batter', '')
        pid       = player_id_map.get(batter)
        score_val = float(row.get(rank_col, 0) or 0)

        if not pid or pid not in log_map:
            continue

        info       = log_map[pid]
        hrr_games  = info['hrr_games']
        total_g    = info['total_games']
        hrr_rate   = info['hrr_rate']
        avg_h      = info['avg_h']
        avg_r      = info['avg_r']
        avg_rbi    = info['avg_rbi']

        # Color by rate
        if hrr_rate >= 0.60:
            rate_col, rate_bg = "#4ade80", "#052E16"
            rate_lbl = "🔥 HOT"
        elif hrr_rate >= 0.40:
            rate_col, rate_bg = "#fbbf24", "#1C1500"
            rate_lbl = "✅ GOOD"
        else:
            rate_col, rate_bg = "#f87171", "#1C0000"
            rate_lbl = "❄️ COLD"

        tier_lbl, _ = _score_tier(score_val)

        cards_html += f"""
        <div style="background:#0f1923;border:1px solid #1e2d3d;border-radius:10px;
             padding:.65rem .75rem;position:relative">
          <div style="font-size:.65rem;color:#64748b;text-transform:uppercase;
               letter-spacing:.06em;margin-bottom:.2rem">{tier_lbl} · {score_val:.1f}</div>
          <div style="font-size:.92rem;font-weight:700;color:#e2e8f0;
               white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{batter}</div>
          <div style="font-size:.7rem;color:#64748b;margin:.15rem 0 .4rem">
               {row.get('Team','')} vs {row.get('Pitcher','')}</div>
          <div style="background:{rate_bg};border-radius:6px;padding:.3rem .5rem;
               text-align:center;margin-bottom:.35rem">
            <span style="font-size:.78rem;font-weight:700;color:{rate_col}">{rate_lbl}</span>
            <span style="font-size:.72rem;color:{rate_col};margin-left:.4rem;
                  font-family:'JetBrains Mono',monospace">
              {hrr_games}/{total_g}G with H+R+RBI ≥2
            </span>
          </div>
          <div style="display:flex;gap:.4rem;font-size:.68rem;color:#94a3b8;
               font-family:'JetBrains Mono',monospace">
            <span>H {avg_h:.2f}/G</span>
            <span>R {avg_r:.2f}/G</span>
            <span>RBI {avg_rbi:.2f}/G</span>
          </div>
        </div>"""

    cards_html += '</div>'
    st.markdown(cards_html, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# PLAYER DEEP DIVE
# ─────────────────────────────────────────────────────────────────────────────

def render_player_deep_dive(filtered_df: pd.DataFrame, player_id_map: dict):
    if filtered_df.empty:
        return

    with st.expander("🔬 Player Deep Dive — Game Log & Statcast", expanded=False):
        st.markdown(
            '<div class="notice notice-info">Select any player from the filtered list '
            'to see their last 10 games + rolling quality-of-contact metrics.</div>',
            unsafe_allow_html=True
        )
        players    = sorted(filtered_df['Batter'].unique().tolist())
        sel_player = st.selectbox("Select Player", players, key="deep_dive_player")
        pid        = player_id_map.get(sel_player)

        col1, col2 = st.columns([3, 2])

        with col1:
            st.markdown(f"**📅 {sel_player} — Last 10 Games**")
            if pid:
                from mlb_api import get_player_game_log
                log_df = get_player_game_log(pid, last_n=10)
                if not log_df.empty:
                    def _style_hits(val):
                        return 'color:#10b981;font-weight:700' if val > 0 else ''
                    styled_log = (
                        log_df.style
                        .map(_style_hits, subset=['H'])
                        .format({'AVG': '{}'})
                    )
                    st.dataframe(styled_log, width='stretch', hide_index=True)

                    if 'H' in log_df.columns and 'AB' in log_df.columns:
                        played   = len(log_df[log_df['AB'] > 0])
                        hit_games= len(log_df[log_df['H'] > 0])
                        total_h  = log_df['H'].sum()
                        total_hr = log_df['HR'].sum()
                        rate     = hit_games / played if played else 0
                        color    = ("var(--hit)" if rate >= 0.6 else
                                    "var(--xb)"  if rate >= 0.4 else "var(--hr)")
                        tier, tc = _score_tier(rate * 100)
                        st.markdown(
                            f'<div style="font-size:.78rem;color:var(--muted);margin-top:.4rem">'
                            f'Last {played} G: <span style="color:{color};font-weight:700">'
                            f'{hit_games} hits in {played} games</span> · '
                            f'{total_h} H · {total_hr} HR &nbsp;'
                            f'<span style="background:{tc}22;color:{tc};padding:1px 6px;'
                            f'border-radius:20px;font-size:.65rem;font-weight:700">{tier}</span>'
                            f'</div>',
                            unsafe_allow_html=True
                        )
                else:
                    st.info("No game log data available yet for this season.")
            else:
                st.info("⏳ Player ID not resolved — game log unavailable.")

        with col2:
            st.markdown(f"**⚡ Statcast — Last 30 Days**")
            if pid:
                try:
                    from savant import get_batter_quality_metrics
                    metrics = get_batter_quality_metrics(pid, days_back=30)
                    if metrics:
                        metric_rows = [
                            ("Sample",    f"{metrics.get('sample_size','—')} BIP",  "var(--muted)"),
                            ("Avg EV",    f"{metrics.get('avg_ev','—')} mph",
                             "var(--hit)" if metrics.get('avg_ev',0) >= 90 else "var(--text)"),
                            ("Max EV",    f"{metrics.get('max_ev','—')} mph",
                             "var(--hit)" if metrics.get('max_ev',0) >= 105 else "var(--text)"),
                            ("Barrel%",   f"{metrics.get('barrel_pct','—')}%",
                             "var(--hit)" if metrics.get('barrel_pct',0) >= 8 else "var(--text)"),
                            ("Hard Hit%", f"{metrics.get('hard_hit_pct','—')}%",
                             "var(--hit)" if metrics.get('hard_hit_pct',0) >= 40 else "var(--text)"),
                            ("Avg LA",    f"{metrics.get('avg_la','—')}°",    "var(--text)"),
                            ("xBA",       f"{metrics.get('xba','—')}",        "var(--text)"),
                            ("xwOBA",     f"{metrics.get('xwoba','—')}",      "var(--text)"),
                        ]
                        rows_html = "".join(
                            f'<div class="pcard-row">'
                            f'<span class="pk">{label}</span>'
                            f'<span class="pv" style="color:{color}">{val}</span>'
                            f'</div>'
                            for label, val, color in metric_rows
                        )
                        st.markdown(
                            f'<div class="pcard" style="margin-top:0">{rows_html}</div>',
                            unsafe_allow_html=True
                        )
                    else:
                        st.info(f"No Statcast data in last 30 days.")
                except Exception:
                    st.info("Statcast unavailable.")
            else:
                st.info("⏳ Player ID not resolved.")


# ─────────────────────────────────────────────────────────────────────────────
# BEST PER TARGET
# ─────────────────────────────────────────────────────────────────────────────

def render_best_per_target(slate_df: pd.DataFrame, filters: dict):
    if len(slate_df) < 3:
        return

    with st.expander("🔍 Best Per Target — Full Slate", expanded=True):
        st.markdown(
            '<div class="notice notice-info">ℹ️ Full slate (exclusions only). '
            'Not affected by target/stat filters.</div>',
            unsafe_allow_html=True
        )

        defs = [
            ('Hit_Score',    'pcard-hit',    '🎯', 'HIT',      'Any Base Hit'),
            ('Single_Score', 'pcard-single', '1️⃣', 'SINGLE',  'Single'),
            ('XB_Score',     'pcard-xb',     '🔥', 'XB',       'Double / Triple'),
            ('HR_Score',     'pcard-hr',     '💣', 'HR',        'Home Run'),
            ('HRR_Score',    'pcard-hit',    '🔴', 'H+R+RBI',  'Hits+Runs+RBIs'),
        ]

        use_gc     = filters.get('use_gc', False)
        cards_html = '<div class="pcard-grid">'
        LG         = CONFIG

        for sc, css, icon, short, desc in defs:
            if sc not in slate_df.columns:
                continue
            rank_sc  = (sc + '_gc') if (use_gc and sc + '_gc' in slate_df.columns) else sc
            eligible = _eligible_for_target(slate_df, sc)
            row      = eligible.loc[eligible[rank_sc].idxmax()]
            disp_val = float(row[rank_sc])
            base_col = sc + '_base'

            tier_lbl, tier_color = _score_tier(disp_val)
            profile_badge = _profile_badge(row, sc)

            park_row = ""
            if filters['use_park'] and base_col in slate_df.columns \
                    and row.get(base_col, 0) != 0:
                delta    = row[sc] - row[base_col]
                pct      = delta / row[base_col] * 100
                col_     = "var(--pos)" if delta >= 0 else "var(--neg)"
                park_row = (
                    f'<div class="pcard-row"><span class="pk">Park Δ</span>'
                    f'<span class="pv" style="color:{col_}">'
                    f'{("+" if delta>=0 else "")}{pct:.1f}%</span></div>'
                )

            k_lg  = LG['league_k_avg']  - row['p_k']
            bb_lg = LG['league_bb_avg'] - row['p_bb']
            hr_lg = row['p_hr']         - LG['league_hr_avg']
            k_cls  = "pos-val" if k_lg  >= 0 else "neg-val"
            bb_cls = "pos-val" if bb_lg >= 0 else "neg-val"
            hr_cls = "pos-val" if hr_lg >= 0 else "neg-val"
            gph    = grade_pill(str(row.get('pitch_grade', 'B')))

            hist_row = ""
            if row['PA'] >= LG['hist_min_pa']:
                hist_row = (
                    f'<div class="pcard-row"><span class="pk">Hist PA</span>'
                    f'<span class="pv">{int(row["PA"])} PA · {row["AVG"]:.3f}</span></div>'
                )

            sc_row = ""
            if 'Barrel%' in slate_df.columns and pd.notna(row.get('Barrel%')):
                sc_row = (
                    f'<div class="pcard-row"><span class="pk">Barrel%</span>'
                    f'<span class="pv" style="color:var(--hit)">'
                    f'{row["Barrel%"]:.1f}%</span></div>'
                )

            # Batting order slot
            order_row = ""
            if '_order_pos' in slate_df.columns and pd.notna(row.get('_order_pos')):
                try:
                    slot = int(float(row['_order_pos']))   # float() first handles "4.0" strings
                except (ValueError, TypeError):
                    slot = None
                if slot:
                    slot_ctx   = {1:"Leadoff",2:"#2 Hitter",3:"#3 Hitter",
                                  4:"Cleanup",5:"#5 Hitter"}.get(slot, f"#{slot} Hitter")
                    slot_color = "var(--hit)" if slot in (3,4,5) else \
                                 "var(--accent)" if slot in (1,2) else "var(--muted)"
                    order_row  = (
                        f'<div class="pcard-row"><span class="pk">Lineup Slot</span>'
                        f'<span class="pv" style="color:{slot_color}">#{slot} — {slot_ctx}</span></div>'
                    )

            # Rolling form badge
            form_row = ""
            if '_form_label' in slate_df.columns and pd.notna(row.get('_form_label')):
                flabel = row['_form_label']
                frate  = row.get('_form_rate')
                fcolor = "var(--hit)" if '🔥' in str(flabel) else "var(--hr)"
                frate_str = f" ({frate:.2f} H/G)" if pd.notna(frate) else ""
                form_row = (
                    f'<div class="pcard-row"><span class="pk">7-Day Form</span>'
                    f'<span class="pv" style="color:{fcolor}">'
                    f'{flabel}{frate_str}</span></div>'
                )

            # Profile gap explanation row — shows the gap that triggered the badge
            profile_row = ""
            if sc == 'Single_Score':
                xb_val  = float(row.get('XB_Score', 0) or 0)
                hr_val  = float(row.get('HR_Score',  0) or 0)
                xb_gap  = xb_val - disp_val
                hr_gap  = hr_val - disp_val
                if xb_gap >= 7:
                    profile_row = (
                        f'<div class="pcard-row"><span class="pk">XB vs Single gap</span>'
                        f'<span class="pv" style="color:#f59e0b">'
                        f'XB={xb_val:.1f} (+{xb_gap:.1f})</span></div>'
                    )
                elif hr_gap >= 12:
                    profile_row = (
                        f'<div class="pcard-row"><span class="pk">HR vs Single gap</span>'
                        f'<span class="pv" style="color:#ef4444">'
                        f'HR={hr_val:.1f} (+{hr_gap:.1f})</span></div>'
                    )
            elif sc == 'XB_Score':
                hr_val = float(row.get('HR_Score', 0) or 0)
                hr_gap = hr_val - disp_val
                if hr_gap >= 12:
                    profile_row = (
                        f'<div class="pcard-row"><span class="pk">HR vs XB gap</span>'
                        f'<span class="pv" style="color:#ef4444">'
                        f'HR={hr_val:.1f} (+{hr_gap:.1f})</span></div>'
                    )

            cards_html += f"""
            <div class="pcard {css}">
              <div class="pcard-header">
                <div>
                  <div class="pcard-name">{row['Batter']}{profile_badge}</div>
                  <div class="pcard-team">{row['Team']} · {icon} {desc}</div>
                </div>
                <div style="text-align:right">
                  <div class="pcard-score">{disp_val:.1f}</div>
                  <div style="font-size:.6rem;padding:1px 5px;border-radius:20px;
                    background:{tier_color}22;color:{tier_color};font-weight:700;
                    margin-top:.2rem;font-family:'JetBrains Mono',monospace">{tier_lbl}</div>
                </div>
              </div>
              <div class="pcard-row"><span class="pk">Pitcher</span>
                <span class="pv">{row['Pitcher']} {gph}</span></div>
              <div class="pcard-row"><span class="pk">Hit Prob</span>
                <span class="pv">{row['total_hit_prob']:.1f}%</span></div>
              <div class="pcard-row"><span class="pk">1B / XB / HR</span>
                <span class="pv">{row['p_1b']:.1f} / {row['p_xb']:.1f} / {row['p_hr']:.1f}%</span></div>
              <div class="pcard-row"><span class="pk">K%</span>
                <span class="pv {k_cls}">{row['p_k']:.1f}% ({k_lg:+.1f} vs lg)</span></div>
              <div class="pcard-row"><span class="pk">BB%</span>
                <span class="pv {bb_cls}">{row['p_bb']:.1f}% ({bb_lg:+.1f} vs lg)</span></div>
              <div class="pcard-row"><span class="pk">HR vs Lg</span>
                <span class="pv {hr_cls}">{hr_lg:+.2f}%</span></div>
              <div class="pcard-row"><span class="pk">vs Grade</span>
                <span class="pv">{int(row['vs Grade'])}</span></div>
              {profile_row}{order_row}{form_row}{sc_row}{park_row}{hist_row}
            </div>"""

        cards_html += '</div>'
        st.markdown(cards_html, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# VISUALIZATIONS
# ─────────────────────────────────────────────────────────────────────────────

def render_visualizations(df: pd.DataFrame, filtered_df: pd.DataFrame, score_col: str):
    with st.expander("📈 Charts & Team Summary", expanded=False):
        axis_cfg  = alt.Axis(gridColor='#1e2d3d', domainColor='#1e2d3d',
                             labelColor='#64748b', titleColor='#64748b', labelFontSize=9)
        title_cfg = lambda t: alt.TitleParams(t, color='#94a3b8', fontSize=11)

        c1, c2 = st.columns(2)
        with c1:
            ch = alt.Chart(df).mark_bar(
                color='#3b82f6', opacity=0.8, cornerRadiusTopLeft=3, cornerRadiusTopRight=3
            ).encode(
                alt.X(f'{score_col}:Q', bin=alt.Bin(maxbins=15), title='Score', axis=axis_cfg),
                alt.Y('count()', title='Players', axis=axis_cfg),
                tooltip=['count()']
            ).properties(title=title_cfg(f'{score_col} Distribution'), width=300, height=200)
            st.altair_chart(ch.configure_view(strokeWidth=0), width="stretch")

        with c2:
            if not filtered_df.empty:
                ch2 = alt.Chart(filtered_df).mark_circle(size=80, opacity=0.85).encode(
                    alt.X('total_hit_prob:Q', title='Hit Prob %', axis=axis_cfg),
                    alt.Y('p_k:Q',            title='K Prob %',   axis=axis_cfg),
                    alt.Color(f'{score_col}:Q', scale=alt.Scale(scheme='viridis'), legend=None),
                    alt.Size('p_hr:Q', legend=None),
                    tooltip=['Batter','Team',
                             alt.Tooltip(score_col, format='.1f'),
                             'total_hit_prob','p_k','p_hr','pitch_grade']
                ).properties(title=title_cfg('Hit Prob vs K Risk'), width=300, height=200)
                st.altair_chart(ch2.configure_view(strokeWidth=0), width="stretch")

        if not filtered_df.empty and len(filtered_df) <= 30:
            st.markdown("**Individual Score Breakdowns**")
            score_defs = [
                ('Hit_Score',    '#10b981', '🎯 Hit Score'),
                ('Single_Score', '#06b6d4', '1️⃣ Single Score'),
                ('XB_Score',     '#f59e0b', '🔥 XB Score'),
                ('HR_Score',     '#ef4444', '💣 HR Score'),
            ]
            r1c1, r1c2 = st.columns(2)
            r2c1, r2c2 = st.columns(2)
            for i, (sc, colour, label) in enumerate(score_defs):
                if sc not in filtered_df.columns:
                    continue
                chart_df = filtered_df[['Batter', sc]].sort_values(sc, ascending=False)
                ch_s = alt.Chart(chart_df).mark_bar(
                    color=colour, opacity=0.85, cornerRadiusTopLeft=2, cornerRadiusTopRight=2
                ).encode(
                    alt.X('Batter:N', sort='-y',
                          axis=alt.Axis(labelAngle=-45, labelFontSize=8,
                                        labelColor='#64748b', domainColor='#1e2d3d')),
                    alt.Y(f'{sc}:Q', scale=alt.Scale(domain=[0, 100]),
                          axis=alt.Axis(labelFontSize=8, labelColor='#64748b',
                                        domainColor='#1e2d3d', gridColor='#1e2d3d',
                                        title='Score')),
                    tooltip=['Batter', alt.Tooltip(f'{sc}:Q', format='.1f', title='Score')]
                ).properties(title=title_cfg(label), width=250, height=180)
                with [r1c1, r1c2, r2c1, r2c2][i]:
                    st.altair_chart(ch_s.configure_view(strokeWidth=0), width="stretch")

        if not filtered_df.empty:
            ts = filtered_df.groupby('Team').agg(
                Players    =('Batter',        'count'),
                AvgHitProb =('total_hit_prob', 'mean'),
                AvgHit     =('Hit_Score',       'mean'),
                AvgXB      =('XB_Score',        'mean'),
                AvgHR      =('HR_Score',        'mean'),
            ).round(1).sort_values('AvgHitProb', ascending=False).reset_index()
            ts.columns = ['Team','Players','Avg Hit%','🎯 Hit','🔥 XB','💣 HR']
            st.markdown("**Team Summary**")
            st.dataframe(ts, width='stretch', hide_index=True)
