"""
unders.py — Under Targets Page
================================
Flips the predictor logic: instead of finding who WILL get hits/XBs/HRs,
finds players unlikely to accumulate bases — useful for total bases unders,
extra base hit unders, and no-hit props.

Architecture mirrors the main Predictor page:
  - Same sidebar filters (starters, confirmed lineups, team, exclusions)
  - Same df (already scored by engine.py)
  - New compute_under_scores() adds Under_Score columns
  - New rendering logic inverts the table (low is good)

Three under target types:
  XB Under   — player unlikely to get an extra base hit (double/triple)
  TB Under   — player unlikely to accumulate total bases (line at 1.5 or 2.0)
  Hit Under  — player unlikely to get any base hit (0.5 line, highest variance)

Disqualification logic:
  Each target has offsetting categories that can STILL produce bases even
  when the primary category is low. Any player exceeding the disqualification
  threshold in an offsetting category is flagged ⚠️ and ranked lower.
  Users can override with the "Show disqualified" toggle.

Under_Score computation (per type):
  XB Under:  (100 - XB_Score)*0.5 + (100 - HR_Score)*0.35
             + (100 - Hit_Score)*0.15 + K_bonus + pitcher_bonus
  TB Under:  (100 - Hit_Score)*0.45 + (100 - XB_Score)*0.35
             + (100 - HR_Score)*0.20 + K_bonus + pitcher_bonus
  Hit Under: (100 - Hit_Score)*0.70 + K_bonus*1.5 + pitcher_bonus
             (K% matters more here — strikeout = guaranteed no hit)

Higher Under_Score = better under candidate.
"""

import streamlit as st
import pandas as pd
import numpy as np
from config import CONFIG
from helpers import grade_pill, style_grade_cell


# ─────────────────────────────────────────────────────────────────────────────
# UNDER SCORE COMPUTATION
# ─────────────────────────────────────────────────────────────────────────────

def compute_under_scores(df: pd.DataFrame,
                         form_map:     dict | None = None,
                         use_gc:       bool = True,
                         pitcher_form: dict | None = None,
                         rest_map:     dict | None = None) -> pd.DataFrame:
    """
    Full 8+1+1+1 layer under score computation. Vectorized throughout.

    Signal architecture:
      Layer 1  — Primary offensive scores, GC-adjusted when use_gc=True
      Layer 2  — K% and BB%
      Layer 3  — Pitcher grade + recent ERA/WHIP (7-day) + days rest / last IP
      Layer 4  — Historical matchup AVG/PA
      Layer 5  — Recent XB rate (7-day)
      Layer 6  — Recent hit rate (7-day cold hitter signal)
      Layer 7  — Statcast contact quality
      Layer 8  — vs Grade + park factor
      Layer 9  — Game conditions suppression
      Layer 10 — Batting order position
      Layer 11 — Platoon split

    rest_map: {pitcher_name: {days_rest, last_ip, rest_signal}} from get_pitcher_rest_map()
    """
    df           = df.copy()
    cfg          = CONFIG
    form_map     = form_map     or {}
    pitcher_form = pitcher_form or {}
    rest_map     = rest_map     or {}

    for col in ['Hit_Score','Single_Score','XB_Score','HR_Score']:
        if col not in df.columns:
            df[col] = 50.0

    # ── FIX 1: GC-adjusted scores as Layer 1 ─────────────────────────────────
    def _gc_or_base(gc_col: str, base_col: str) -> pd.Series:
        if use_gc and gc_col in df.columns:
            return pd.to_numeric(df[gc_col], errors='coerce').fillna(
                pd.to_numeric(df[base_col], errors='coerce').fillna(50.0)
            ).clip(0, 100)
        return df[base_col].clip(0, 100)

    hit    = _gc_or_base('Hit_Score_gc',    'Hit_Score')
    xb     = _gc_or_base('XB_Score_gc',     'XB_Score')
    hr     = _gc_or_base('HR_Score_gc',     'HR_Score')
    single = df['Single_Score'].clip(0, 100)

    k  = pd.to_numeric(df.get('p_k',  0), errors='coerce').fillna(0)
    bb = pd.to_numeric(df.get('p_bb', 0), errors='coerce').fillna(0)

    # ── Layer 2: K% and BB% ───────────────────────────────────────────────────
    k_bonus = ((k - cfg['league_k_avg']) * cfg['under_k_weight']).clip(lower=0)

    bb_xb   = ((bb - cfg['league_bb_avg']) * cfg['under_bb_weight_xb']  ).clip(lower=0)
    bb_tb15 = ((bb - cfg['league_bb_avg']) * cfg['under_bb_weight_tb15']).clip(lower=0)
    bb_tb05 = ((bb - cfg['league_bb_avg']) * cfg['under_bb_weight_tb05']).clip(lower=0)
    bb_hit  = ((bb - cfg['league_bb_avg']) * cfg['under_bb_weight_hit'] ).clip(lower=0)

    # ── Layer 3: Pitcher grade + recent form ──────────────────────────────────
    # Grade bonus: static season-level quality
    def _pgbonus(grade):
        return (cfg['under_pitcher_bonus'] if grade == 'A+' else
                cfg['under_pitcher_a']     if grade == 'A'  else 0.0)

    p_bonus = df['pitch_grade'].apply(_pgbonus) \
              if 'pitch_grade' in df.columns \
              else pd.Series(0.0, index=df.index)

    # Pitcher recent form bonus: low ERA/WHIP in last 7 days = extra suppression
    # ERA league avg ~4.20 — below = dominant, above = hittable
    # WHIP league avg ~1.30 — below = clean, above = hitters reaching base
    ERA_LG  = 4.20
    WHIP_LG = 1.30

    def _pitcher_form_bonus(pitcher_name: str) -> float:
        if not pitcher_name or not pitcher_form:
            return 0.0
        # Try full name then last name
        pf = pitcher_form.get(pitcher_name)
        if not pf:
            last = pitcher_name.split()[-1]
            pf   = pitcher_form.get(last)
        if not pf or pf.get('games', 0) < 1:
            return 0.0
        era_bonus  = np.clip((ERA_LG  - pf['era'])  * 0.8, -3.0, 3.0)
        whip_bonus = np.clip((WHIP_LG - pf['whip']) * 3.0, -2.0, 2.0)
        return float(era_bonus + whip_bonus)

    p_form_bonus = df['Pitcher'].apply(_pitcher_form_bonus) \
                   if 'Pitcher' in df.columns \
                   else pd.Series(0.0, index=df.index)

    # Rest/workload signal — days since last start and IP from that outing
    # Well-rested pitcher = sharper arm = better for unders
    # Short rest / early exit last time = fatigue/uncertainty = worse for unders
    def _rest_signal(pitcher_name: str) -> float:
        if not pitcher_name or not rest_map:
            return 0.0
        info = rest_map.get(pitcher_name)
        if not info:
            last = pitcher_name.split()[-1]
            info = rest_map.get(last)
        if not info:
            return 0.0
        return float(info.get('rest_signal', 0.0))

    p_rest_bonus = df['Pitcher'].apply(_rest_signal) \
                   if 'Pitcher' in df.columns \
                   else pd.Series(0.0, index=df.index)

    # ─────────────────────────────────────────────────────────────────────────
    # LAYER 4 — Historical matchup AVG/PA
    # ─────────────────────────────────────────────────────────────────────────
    hist_min_pa = cfg['under_hist_min_pa']
    hist_weight = cfg['under_hist_weight']
    hist_max    = cfg['under_hist_max_adj']
    league_avg  = cfg['league_avg']

    pa_col  = pd.to_numeric(df.get('PA',  0), errors='coerce').fillna(0)
    avg_col = pd.to_numeric(df.get('AVG', 0), errors='coerce').fillna(0)

    hist_raw  = (league_avg - avg_col) * hist_weight
    hist_base = hist_raw.where(pa_col >= hist_min_pa, 0.0).clip(-hist_max, hist_max)

    hist_xb   = hist_base * 0.6
    hist_tb15 = hist_base * 0.8
    hist_tb05 = hist_base * 1.0
    hist_hit  = hist_base * 1.2

    # ─────────────────────────────────────────────────────────────────────────
    # LAYER 5 — Recent XB rate (7-day 2B+3B/G)
    # ─────────────────────────────────────────────────────────────────────────
    xb_rate_lg  = cfg['under_xb_rate_lg_avg']
    xb_rate_w   = cfg['under_xb_rate_weight']
    xb_rate_max = cfg['under_xb_rate_max_adj']

    def _xb_rate_sig(batter: str) -> float:
        info = form_map.get(batter, {})
        if not info or info.get('games', 0) < 3:
            return 0.0
        rate = info.get('xb_rate', xb_rate_lg)
        return float(np.clip((xb_rate_lg - rate) * xb_rate_w,
                             -xb_rate_max, xb_rate_max))

    xb_rate_adj = df['Batter'].apply(_xb_rate_sig)

    # ─────────────────────────────────────────────────────────────────────────
    # LAYER 6 — Recent hit rate (7-day H/G)
    # ─────────────────────────────────────────────────────────────────────────
    hit_rate_lg  = cfg['under_hit_rate_lg_avg']
    hit_rate_w   = cfg['under_hit_rate_weight']
    hit_rate_max = cfg['under_hit_rate_max_adj']

    def _hit_rate_sig(batter: str) -> float:
        info = form_map.get(batter, {})
        if not info or info.get('games', 0) < 3:
            return 0.0
        rate = info.get('hit_rate', hit_rate_lg)
        return float(np.clip((hit_rate_lg - rate) * hit_rate_w,
                             -hit_rate_max, hit_rate_max))

    hit_rate_adj = df['Batter'].apply(_hit_rate_sig)

    # ─────────────────────────────────────────────────────────────────────────
    # LAYER 7 — Statcast contact quality
    # ─────────────────────────────────────────────────────────────────────────
    def _sc(col: str, lg_val: float) -> pd.Series:
        if col not in df.columns:
            return pd.Series(lg_val, index=df.index)
        return pd.to_numeric(df[col], errors='coerce').fillna(lg_val)

    barrel = _sc('Barrel%', cfg['league_barrel_pct'])
    hh     = _sc('HH%',     cfg['league_hh_pct'])
    avgev  = _sc('AvgEV',   cfg['league_avgev'])
    xslg   = _sc('xSLG',    cfg['league_xslg'])
    xba    = _sc('xBA',     cfg['league_xba'])
    xwoba  = _sc('xwOBA',   cfg['league_xwoba'])

    barrel_adj = ((cfg['league_barrel_pct'] - barrel) * cfg['under_barrel_weight']
                  ).clip(-cfg['under_barrel_max'], cfg['under_barrel_max'])
    hh_adj     = ((cfg['league_hh_pct'] - hh) * cfg['under_hh_weight']
                  ).clip(-cfg['under_hh_max'], cfg['under_hh_max'])
    avgev_adj  = ((cfg['league_avgev'] - avgev) * cfg['under_avgev_weight']
                  ).clip(-cfg['under_avgev_max'], cfg['under_avgev_max'])
    xslg_adj   = ((cfg['league_xslg'] - xslg) * cfg['under_xslg_weight']
                  ).clip(-cfg['under_xslg_max'], cfg['under_xslg_max'])
    xba_adj    = ((cfg['league_xba']   - xba)   * cfg['under_xba_weight']
                  ).clip(-cfg['under_xba_max'], cfg['under_xba_max'])
    xwoba_adj  = ((cfg['league_xwoba'] - xwoba) * cfg['under_xwoba_weight']
                  ).clip(-cfg['under_xwoba_max'], cfg['under_xwoba_max'])

    # ─────────────────────────────────────────────────────────────────────────
    # LAYER 8 — vs Grade + park XB factor
    # ─────────────────────────────────────────────────────────────────────────
    vs_grade   = pd.to_numeric(df.get('vs Grade', 0), errors='coerce').fillna(0)
    vsgrade_adj = ((-vs_grade).clip(lower=0) * cfg['under_vsgrade_weight']
                   ).clip(0, cfg['under_vsgrade_max'])

    xb_boost  = pd.to_numeric(df.get('xb_boost', 0), errors='coerce').fillna(0)
    parkxb_pen = (xb_boost * cfg['under_parkxb_weight']
                  ).clip(0, cfg['under_parkxb_max'])

    # ─────────────────────────────────────────────────────────────────────────
    # LAYER 9 — FIX 2: Game conditions suppression signal
    # ─────────────────────────────────────────────────────────────────────────
    # Computes a per-row GC suppression bonus based on today's game environment.
    # Positive = pitcher-friendly conditions = good for unders.
    # Negative = hitter-friendly conditions = bad for unders.
    # Only fires when GC columns are present; gracefully returns 0 otherwise.
    gc_max = cfg['under_gc_max']

    _GC_COLS = ['gc_hits20','gc_runs10','gc_k20','gc_qs','gc_hr4']
    has_gc_data = use_gc and all(c in df.columns for c in _GC_COLS)

    if has_gc_data:
        gc_hits = pd.to_numeric(df['gc_hits20'], errors='coerce').fillna(cfg['gc_hits20_anchor'])
        gc_runs = pd.to_numeric(df['gc_runs10'], errors='coerce').fillna(cfg['gc_runs10_anchor'])
        gc_k    = pd.to_numeric(df['gc_k20'],    errors='coerce').fillna(cfg['gc_k20_anchor'])
        gc_qs   = pd.to_numeric(df['gc_qs'],     errors='coerce').fillna(cfg['gc_qs_anchor'])
        gc_hr4  = pd.to_numeric(df['gc_hr4'],    errors='coerce').fillna(cfg['gc_hr4_anchor'])

        hits_sig = (cfg['gc_hits20_anchor'] - gc_hits) * cfg['under_gc_hits_weight']
        runs_sig = (cfg['gc_runs10_anchor'] - gc_runs) * cfg['under_gc_runs_weight']
        k_sig    = (gc_k  - cfg['gc_k20_anchor']) * cfg['under_gc_k_weight']
        qs_sig   = (gc_qs - cfg['gc_qs_anchor'])  * cfg['under_gc_qs_weight']
        hr_sig   = (cfg['gc_hr4_anchor'] - gc_hr4) * cfg['under_gc_hr_weight']

        gc_general    = (hits_sig + runs_sig + k_sig + qs_sig).clip(-gc_max, gc_max)
        gc_power_supp = (hits_sig + runs_sig + k_sig + qs_sig + hr_sig).clip(-gc_max, gc_max)
        # HRR under: low-scoring environment = fewer run/RBI opportunities
        # Invert runs_sig — below-median runs = good for HRR under = positive
        gc_hrr_over   = (runs_sig * 1.5 + hits_sig * 0.5).clip(-gc_max, gc_max)
    else:
        gc_general    = pd.Series(0.0, index=df.index)
        gc_power_supp = pd.Series(0.0, index=df.index)
        gc_hrr_over   = pd.Series(0.0, index=df.index)

    # ─────────────────────────────────────────────────────────────────────────
    # LAYER 10 — Batting order position
    # ─────────────────────────────────────────────────────────────────────────
    # Late-order batters (8-9) get fewer quality at-bats, face fresh pitchers
    # more often, and are typically weaker offensive players.
    # → Small bonus for under targets (they're less likely to accumulate bases).
    # Cleanup hitters (3-5) see the most pitches with runners on, pitchers work
    # carefully, more walks — neutral to slight penalty for under targets.
    # Applied uniformly to all under types, small cap (±2 pts max).
    order_adj = pd.Series(0.0, index=df.index)
    if '_order_pos' in df.columns:
        pos = pd.to_numeric(df['_order_pos'], errors='coerce')
        # Slots 8-9: typically weakest bats → under bonus
        order_adj = order_adj.where(~(pos.isin([8, 9])), order_adj + 1.5)
        # Slots 1-2: table-setters, high OBP, walk-prone → mild under bonus
        order_adj = order_adj.where(~(pos.isin([1, 2])), order_adj + 0.5)
        # Slots 3-5: cleanup, most dangerous offensive spots → slight under penalty
        order_adj = order_adj.where(~(pos.isin([3, 4, 5])), order_adj - 1.0)
        order_adj = order_adj.clip(-2.0, 2.0)

    # ─────────────────────────────────────────────────────────────────────────
    # LAYER 11 — Platoon split
    # ─────────────────────────────────────────────────────────────────────────
    # Same-hand matchups (RHB vs RHP or LHB vs LHP) suppress batting averages
    # historically. Opposite hand = platoon advantage = batter sees ball better.
    # For unders: same-hand = bonus (batter disadvantaged = fewer hits/XBs).
    # Only fires when both pitcher hand and batter hand are known.
    # Batter hand data wires in when MLB roster lookup activates.
    # Conservative: pitcher hand is known (from mlb_api), batter hand not yet
    # programmatically fetched — so we apply a mild pitcher-hand-only signal:
    # LHP facing the lineup = generally favorable for under (LHPs suppress AVG
    # across the slate because most batters are RHH and see fewer LHPs).
    platoon_adj = pd.Series(0.0, index=df.index)
    if '_pitcher_hand' in df.columns:
        # LHP bonus: league-wide RHH AVG vs LHP is ~.010 below vs RHP
        lhp_mask     = df['_pitcher_hand'] == 'L'
        platoon_adj  = platoon_adj.where(~lhp_mask, platoon_adj + 0.8)
        platoon_adj  = platoon_adj.clip(-1.5, 1.5)

    # ─────────────────────────────────────────────────────────────────────────
    # FINAL SCORES — all 9 layers
    # ─────────────────────────────────────────────────────────────────────────

    # ── XB Under ─────────────────────────────────────────────────────────────
    df['Under_XB_Score'] = (
        (100 - xb)  * 0.42
      + (100 - hr)  * 0.22
      + (100 - hit) * 0.08
      + k_bonus + bb_xb
      + p_bonus + p_form_bonus + p_rest_bonus   # L3 — grade + form + rest
      + hist_xb
      + xb_rate_adj
      + barrel_adj + hh_adj + avgev_adj + xslg_adj
      + vsgrade_adj - parkxb_pen
      + gc_power_supp
      + order_adj + platoon_adj
    ).clip(0, 100).round(1)

    # ── TB Under 1.5 ─────────────────────────────────────────────────────────
    single_profile_bonus = ((single - xb) * 0.15).clip(lower=0, upper=5)
    df['Under_TB15_Score'] = (
        (100 - xb)  * 0.45
      + (100 - hr)  * 0.28
      + single_profile_bonus
      + k_bonus + bb_tb15
      + p_bonus + p_form_bonus + p_rest_bonus   # L3
      + hist_tb15
      + xb_rate_adj * 0.7
      + xslg_adj * 0.8 + barrel_adj * 0.5
      + vsgrade_adj - parkxb_pen * 0.5
      + gc_power_supp * 0.8
      + order_adj + platoon_adj
    ).clip(0, 100).round(1)

    # ── TB Under 0.5 ─────────────────────────────────────────────────────────
    df['Under_TB05_Score'] = (
        (100 - hit) * 0.40
      + (100 - xb)  * 0.28
      + (100 - hr)  * 0.22
      + k_bonus * 1.2 + bb_tb05
      + p_bonus + p_form_bonus + p_rest_bonus   # L3
      + hist_tb05
      + hit_rate_adj
      + xba_adj * 0.7 + xwoba_adj * 0.5
      + vsgrade_adj
      + gc_general
      + order_adj + platoon_adj
    ).clip(0, 100).round(1)

    # ── Hit Under ─────────────────────────────────────────────────────────────
    df['Under_Hit_Score'] = (
        (100 - hit) * 0.45
      + k_bonus * 1.8 + bb_hit
      + p_bonus + p_form_bonus + p_rest_bonus   # L3
      + hist_hit
      + hit_rate_adj * 1.2
      + xba_adj + xwoba_adj * 0.7
      + vsgrade_adj * 1.2
      + gc_general * 1.1
      + order_adj + platoon_adj
    ).clip(0, 100).round(1)

    # ── H+R+RBI Under ─────────────────────────────────────────────────────────
    # Fading a player's composite H+R+RBI total.
    # The prop cashes UNDER when the player accumulates fewer than 2 combined.
    # Key: batting order is the dominant signal INVERTED — slots 7-9 are ideal
    # (rarely score or drive in runs). Cleanup slots 3-5 are BAD for this under
    # (most dangerous for accumulating H+R+RBI). Low-scoring game = good.
    # Under_HRR_Score = inverted HRR_Score (in df if Stage 5 ran) + GC suppression.
    hrr_base = pd.Series(50.0, index=df.index)
    if 'HRR_Score' in df.columns:
        hrr_base = pd.to_numeric(df['HRR_Score'], errors='coerce').fillna(50.0)

    # Order signal is flipped: late-order = GOOD for HRR under
    order_hrr_under = -order_adj   # invert: penalty where over gets bonus

    df['Under_HRR_Score'] = (
        (100 - hrr_base) * 0.55  # primary: invert the HRR over score
      + k_bonus * 1.0             # K% still good (no H contribution)
      + bb_hit                    # BB% good (walk = no RBI if bases empty)
      + p_bonus                   # elite pitcher suppresses all
      + hist_hit  * 0.8           # historical AVG (weak matchup = fewer H)
      + hit_rate_adj              # cold hitter = fewer hits = fewer H+R+RBI
      + xwoba_adj * 0.6           # low offensive quality
      + vsgrade_adj               # poor matchup
      - gc_hrr_over               # low-scoring game environment = good for under
      + order_hrr_under           # late-order batters = fewer R/RBI opportunities
    ).clip(0, 100).round(1)

    # ── Disqualification flags ─────────────────────────────────────────────────
    df['_disq_xb']   = (hr  > cfg['under_xb_disq_hr']) | \
                       (hit > cfg['under_xb_disq_hit'])
    df['_disq_tb15'] = (xb  > cfg['under_xb_disq_hr']) | \
                       (hr  > cfg['under_xb_disq_hr'])
    df['_disq_tb05'] = (hit > cfg['under_tb_disq_any']) | \
                       (xb  > cfg['under_tb_disq_any']) | \
                       (hr  > cfg['under_tb_disq_any'])
    df['_disq_hit']  = hit > cfg['under_hit_disq_hit']
    # HRR Under: disqualify when batting order slots 1-5 AND high Hit_Score
    # (cleanup hitters in good spots almost always accumulate H+R+RBI)
    df['_disq_hrr']  = (hit > 65.0) & (
        df.get('_order_pos', pd.Series(6, index=df.index))
          .apply(lambda x: int(x) in [1,2,3,4,5] if pd.notna(x) else False)
    )

    return df


# ─────────────────────────────────────────────────────────────────────────────
# UNDER TIER LABEL (inverted — lower offensive scores = better)
# ─────────────────────────────────────────────────────────────────────────────

def _under_tier(score: float) -> tuple[str, str]:
    """Return (label, color) for an Under_Score. Higher = better under."""
    if score >= 75: return "ELITE",    "#4ade80"
    if score >= 60: return "STRONG",   "#a3e635"
    if score >= 45: return "GOOD",     "#fbbf24"
    if score >= 30: return "MODERATE", "#fb923c"
    return             "RISKY",    "#f87171"


# ─────────────────────────────────────────────────────────────────────────────
# DISQUALIFICATION REASON STRING
# ─────────────────────────────────────────────────────────────────────────────

def _disq_reason(row: pd.Series, target: str) -> str:
    cfg = CONFIG
    reasons = []
    if target == 'xb':
        hr  = float(row.get('HR_Score',  0) or 0)
        hit = float(row.get('Hit_Score', 0) or 0)
        if hr  > cfg['under_xb_disq_hr']:
            reasons.append(f"HR={hr:.0f} (can still go yard)")
        if hit > cfg['under_xb_disq_hit']:
            reasons.append(f"Hit={hit:.0f} (high contact vol)")
    elif target == 'tb15':
        xb = float(row.get('XB_Score', 0) or 0)
        hr = float(row.get('HR_Score', 0) or 0)
        if xb > cfg['under_xb_disq_hr']:
            reasons.append(f"XB={xb:.0f} (likely extra bases)")
        if hr > cfg['under_xb_disq_hr']:
            reasons.append(f"HR={hr:.0f} (4 bases = loses 1.5)")
    elif target == 'tb05':
        for sc, lbl in [('Hit_Score','Hit'),('XB_Score','XB'),('HR_Score','HR')]:
            val = float(row.get(sc, 0) or 0)
            if val > cfg['under_tb_disq_any']:
                reasons.append(f"{lbl}={val:.0f}")
    elif target == 'hit':
        hit = float(row.get('Hit_Score', 0) or 0)
        if hit > cfg['under_hit_disq_hit']:
            reasons.append(f"Hit={hit:.0f}")
    elif target == 'hrr':
        hit = float(row.get('Hit_Score', 0) or 0)
        pos = row.get('_order_pos')
        if hit > 65.0:
            reasons.append(f"Hit={hit:.0f}")
        if pd.notna(pos) and int(pos) in [1,2,3,4,5]:
            reasons.append(f"Slot #{int(pos)} (high R/RBI opportunity)")
    return ", ".join(reasons)


# ─────────────────────────────────────────────────────────────────────────────
# UNDER SIDEBAR FILTERS
# ─────────────────────────────────────────────────────────────────────────────

def build_under_filters(df: pd.DataFrame) -> dict:
    """
    Sidebar filters for the Under Targets page.
    Mirrors the main Predictor filters — same lineup/team/exclusion controls,
    different target selector (XB / TB / Hit under).
    """
    st.sidebar.title("🔻 Under Target Filters")
    st.sidebar.markdown("---")
    filters = {}

    # ── Under type ────────────────────────────────────────────────────────────
    st.sidebar.markdown("### 🎯 Under Type")
    under_map = {
        "🔻 XB Under — No Extra Bases (doubles/triples)":  "xb",
        "📊 TB Under 1.5 — Under 1.5 Total Bases":         "tb15",
        "📉 TB Under 0.5 — No Bases At All":               "tb05",
        "❌ Hit Under — No Hit (0.5 line)":                 "hit",
        "🔴 H+R+RBI Under — Under 1.5 Hits+Runs+RBIs":    "hrr",
    }
    u_label            = st.sidebar.selectbox("Choose Under Target", list(under_map.keys()))
    filters['under']   = under_map[u_label]
    filters['under_label'] = u_label

    score_col_map = {
        'xb':   'Under_XB_Score',
        'tb15': 'Under_TB15_Score',
        'tb05': 'Under_TB05_Score',
        'hit':  'Under_Hit_Score',
        'hrr':  'Under_HRR_Score',
    }
    disq_col_map = {
        'xb':   '_disq_xb',
        'tb15': '_disq_tb15',
        'tb05': '_disq_tb05',
        'hit':  '_disq_hit',
        'hrr':  '_disq_hrr',
    }
    filters['under_score_col'] = score_col_map[filters['under']]
    filters['under_disq_col']  = disq_col_map[filters['under']]

    # ── Show disqualified players ─────────────────────────────────────────────
    filters['show_disq'] = st.sidebar.toggle(
        "Show disqualified players",
        value=False,
        help="Disqualified = player has an offsetting score high enough to still "
             "accumulate bases through another route. ON = show with ⚠️ warning."
    )

    # ── Park / GC ─────────────────────────────────────────────────────────────
    st.sidebar.markdown("### 🏟️ Park Adjustment")
    filters['use_park'] = st.sidebar.toggle("Include Park Factors", value=True)
    st.sidebar.markdown("### 🌦️ Game Conditions")
    filters['use_gc'] = st.sidebar.toggle("🌦️ Game Conditions Weight", value=True)

    # ── Lineup ────────────────────────────────────────────────────────────────
    st.sidebar.markdown("### 📋 Lineup")
    filters['starters_only']  = st.sidebar.checkbox("Starters only", value=False)
    filters['confirmed_only'] = st.sidebar.toggle(
        "✅ Confirmed lineups only", value=False,
        help="Only show batters in a confirmed batting order."
    )

    # ── Stat filters (inverted for unders) ────────────────────────────────────
    st.sidebar.markdown("### 📊 Filters")
    filters['min_k'] = st.sidebar.slider(
        "Min K Prob % (higher = better under)", 0.0, 50.0, 15.0, 0.5,
        help="Focus on batters facing high strikeout matchups"
    )
    filters['min_bb'] = st.sidebar.slider(
        "Min BB Prob % (higher = more walks = better under)", 0.0, 20.0, 0.0, 0.5,
        help="Walks = 0 bases/hits. High walk rate is one of the strongest under signals."
    )
    filters['max_hit_prob'] = st.sidebar.slider(
        "Max Hit Prob % (lower = better under)", 0.0, 50.0, 35.0, 0.5,
        help="Exclude players with high overall hit probability"
    )
    filters['max_vs'] = st.sidebar.slider(
        "Max vs Grade (lower = weaker matchup)", -10, 10, 5, 1,
        help="Target batters in unfavorable pitcher matchups"
    )

    # ── Team filters ──────────────────────────────────────────────────────────
    st.sidebar.markdown("### 🏟️ Team Filters")
    all_teams = sorted(df['Team'].unique().tolist()) if df is not None else []
    filters['include_teams'] = st.sidebar.multiselect("Include Only Teams", options=all_teams)
    filters['exclude_teams'] = st.sidebar.multiselect("Exclude Teams",       options=all_teams)

    # ── Exclusions ────────────────────────────────────────────────────────────
    st.sidebar.markdown("### 🚫 Player Exclusions")
    if 'excluded_players' not in st.session_state:
        st.session_state.excluded_players = []
    all_players = sorted(df['Batter'].unique().tolist()) if df is not None else []
    excl = st.sidebar.multiselect(
        "Players NOT Playing Today",
        options=all_players,
        default=st.session_state.excluded_players,
        key="under_exclusions",
    )
    st.session_state.excluded_players = excl
    filters['excluded_players'] = excl

    # ── Display ───────────────────────────────────────────────────────────────
    st.sidebar.markdown("### 🔢 Display")
    filters['result_count'] = st.sidebar.selectbox("Show Top N", [5,10,15,20,25,30,"All"], index=2)

    return filters


# ─────────────────────────────────────────────────────────────────────────────
# APPLY UNDER FILTERS
# ─────────────────────────────────────────────────────────────────────────────

def apply_under_filters(df: pd.DataFrame, filters: dict) -> pd.DataFrame:
    """Filter and sort the df for under candidates."""
    if df is None or df.empty:
        return pd.DataFrame()

    out = df.copy()

    # Standard exclusions
    excl = filters.get('excluded_players', [])
    if excl:
        out = out[~out['Batter'].isin(excl)]

    if filters.get('starters_only') and 'Starter' in out.columns:
        out = out[out['Starter'] == 1]

    if filters.get('include_teams'):
        out = out[out['Team'].isin(filters['include_teams'])]
    if filters.get('exclude_teams'):
        out = out[~out['Team'].isin(filters['exclude_teams'])]

    # Under-specific stat filters
    if 'p_k' in out.columns:
        out = out[pd.to_numeric(out['p_k'], errors='coerce').fillna(0) >= filters['min_k']]
    if 'p_bb' in out.columns and filters.get('min_bb', 0) > 0:
        out = out[pd.to_numeric(out['p_bb'], errors='coerce').fillna(0) >= filters['min_bb']]
    if 'total_hit_prob' in out.columns:
        out = out[pd.to_numeric(out['total_hit_prob'], errors='coerce').fillna(100)
                  <= filters['max_hit_prob']]
    if 'vs Grade' in out.columns:
        out = out[pd.to_numeric(out['vs Grade'], errors='coerce').fillna(10)
                  <= filters['max_vs']]

    # Disqualification filter
    disq_col = filters['under_disq_col']
    if not filters.get('show_disq') and disq_col in out.columns:
        out = out[~out[disq_col]]

    # ── Profile-prioritized sort for under targets ────────────────────────────
    # When show_disq=True, disqualified players are visible but always ranked
    # below non-disqualified players — regardless of their Under_Score.
    # This ensures users see the cleanest candidates first.
    sc  = filters['under_score_col']
    disq_col = filters['under_disq_col']
    if sc in out.columns:
        if filters.get('show_disq') and disq_col in out.columns:
            out['_disq_rank'] = out[disq_col].astype(int)   # 0=clean, 1=disqualified
            out = out.sort_values(
                ['_disq_rank', sc],
                ascending=[True, False],
                na_position='last'
            ).drop(columns=['_disq_rank'])
        else:
            out = out.sort_values(sc, ascending=False, na_position='last')

    n = filters['result_count']
    if n != "All":
        out = out.head(int(n))

    return out


# ─────────────────────────────────────────────────────────────────────────────
# RENDER BEST UNDERS (top cards, like Today's Best)
# ─────────────────────────────────────────────────────────────────────────────

def _render_under_top_cards(df: pd.DataFrame, filters: dict):
    """Show top 3 under candidates across the full (exclusion-filtered) slate."""
    if df.empty:
        return

    sc      = filters['under_score_col']
    disq_col= filters['under_disq_col']
    target  = filters['under']

    # Only clean (non-disqualified) players for top cards
    clean = df[~df[disq_col]] if disq_col in df.columns else df
    if clean.empty:
        clean = df  # fallback

    clean = clean.sort_values(sc, ascending=False)

    icons = {'xb': '🔻', 'tb': '📊', 'hit': '❌'}
    icon  = icons.get(target, '🔻')

    cards_html = '<div class="score-grid">'
    for i, (_, row) in enumerate(clean.head(3).iterrows()):
        score     = float(row.get(sc, 0) or 0)
        tier, tc  = _under_tier(score)
        hit_val   = float(row.get('Hit_Score',0) or 0)
        xb_val    = float(row.get('XB_Score', 0) or 0)
        hr_val    = float(row.get('HR_Score', 0) or 0)
        k_val     = float(row.get('p_k',      0) or 0)
        gph       = grade_pill(str(row.get('pitch_grade','B')))

        cards_html += f"""
        <div class="scard scard-hr" style="border-color:#8b5cf6">
          <div class="sc-type" style="color:#8b5cf6">
            <span>{icon} UNDER #{i+1}</span> — {filters['under_label'].split('—')[0].strip()}
          </div>
          <div class="sc-name">{row['Batter']}</div>
          <div class="sc-meta" style="font-size:.7rem">
            {row.get('Team','?')} vs {row.get('Pitcher','?')} {gph}
          </div>
          <div class="sc-score" style="color:#8b5cf6">{score:.1f}</div>
          <div style="margin-top:.4rem">
            <span style="font-size:.62rem;padding:1px 6px;border-radius:20px;
              background:{tc}22;color:{tc};font-weight:700;
              font-family:'JetBrains Mono',monospace">{tier}</span>
          </div>
          <div style="font-size:.68rem;color:#64748b;margin-top:.3rem;line-height:1.5">
            Hit={hit_val:.0f} · XB={xb_val:.0f} · HR={hr_val:.0f} · K%={k_val:.1f}%
          </div>
        </div>"""

    cards_html += '</div>'
    st.markdown(cards_html, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# RENDER UNDER TABLE
# ─────────────────────────────────────────────────────────────────────────────

def _render_under_table(filtered_df: pd.DataFrame, filters: dict):
    if filtered_df.empty:
        st.warning("⚠️ No players match your filters. Try relaxing thresholds.")
        return

    target  = filters['under']
    sc      = filters['under_score_col']
    disq_col= filters['under_disq_col']
    cfg     = CONFIG

    disp = filtered_df.copy()

    # Compute is-disqualified for display (may include them with warning)
    disp['Disq'] = disp[disq_col].map({True: '⚠️', False: '✅'}) \
                   if disq_col in disp.columns else '✅'

    # Rename Under_Score for display
    label_map = {
        'Under_XB_Score':   '🔻 XB Under',
        'Under_TB15_Score': '📊 TB Under 1.5',
        'Under_TB05_Score': '📉 TB Under 0.5',
        'Under_Hit_Score':  '❌ Hit Under',
        'Under_HRR_Score':  '🔴 H+R+RBI Under',
    }

    # Derive display rest column from pitcher name + rest_map
    # Done here (display layer) rather than in compute — keeps scoring clean
    if 'Pitcher' in disp.columns:
        try:
            from mlb_api import get_pitcher_rest_map as _grm
            _rm = _grm()
            def _rest_label(pitcher: str) -> str:
                if not pitcher or not _rm: return "—"
                info = _rm.get(pitcher) or _rm.get(pitcher.split()[-1])
                if not info: return "—"
                d = info.get('days_rest', 0)
                if d >= 5:   return f"✅ {d}d"
                elif d == 4: return f"✅ {d}d"
                elif d == 3: return f"⚠️ {d}d"
                else:        return f"❌ {d}d"
            disp['Rest'] = disp['Pitcher'].apply(_rest_label)
            has_rest = disp['Rest'].ne("—").any()
        except Exception:
            has_rest = False
    else:
        has_rest = False

    col_order = ['Batter','Team','Pitcher','pitch_grade','Rest','Disq',
                 sc,
                 'prop_tb_line','prop_tb_under_odds','prop_tb_over_odds',
                 'prop_hr_odds',
                 'bvp_avg','bvp_ops','bvp_ab',
                 'Hit_Score','XB_Score','HR_Score',
                 'p_k','p_bb','total_hit_prob','p_1b','p_xb','p_hr',
                 'vs Grade',
                 'PA','H','AVG']    # historical matchup data — always last

    rename   = {
        sc:            label_map.get(sc, 'Under Score'),
        'pitch_grade': 'P.Grd ↑',    # ↑ = higher grade = better for under
        'Rest':        'Rest',
        'p_k':         'K%',
        'p_bb':        'BB%',
        'total_hit_prob':'Hit%',
        'p_1b':        '1B%',
        'p_xb':        'XB%',
        'p_hr':        'HR%',
        'vs Grade':    'vsPit',
        'Hit_Score':   '🎯 Hit',
        'XB_Score':    '🔥 XB',
        'HR_Score':    '💣 HR',
        'Disq':        'Status',
        'prop_tb_line':       'TB Line',
        'prop_tb_under_odds': 'TB Under',
        'prop_tb_over_odds':  'TB Over',
        'prop_hr_odds':       'HR Odds',
        'PA':          'Hist PA',
        'H':           'Hist H',
        'AVG':         'Hist AVG',
        'bvp_avg':     'BvP AVG',
        'bvp_ops':     'BvP OPS',
        'bvp_ab':      'BvP AB',
    }

    # Add market edge badge column when prop data available
    has_props = 'prop_tb_under_odds' in disp.columns and \
                disp['prop_tb_under_odds'].astype(bool).any()
    if has_props:
        try:
            from prop_odds import edge_label
            under_score_col = sc
            disp['Market Edge'] = disp.apply(
                lambda r: edge_label(
                    float(r.get(under_score_col, 0) or 0),
                    float(r.get('prop_tb_under_pct', 0) or 0)
                ), axis=1
            )
            col_order_extra = ['Market Edge']
        except Exception:
            col_order_extra = []
    else:
        col_order_extra = []

    existing = [c for c in col_order if c in disp.columns] + \
               [c for c in col_order_extra if c in disp.columns]
    out_df   = disp[existing].rename(columns=rename)

    # ── Under Tier column — insert BEFORE styler is created ───────────────────
    # Must happen before out_df.style.format() — mutating after styler creation
    # causes pandas to crash with a format error on the new string column.
    under_label = label_map.get(sc, 'Under Score')
    if under_label in out_df.columns:
        u_tier = lambda s: (
            "🟢 ELITE"    if s >= 75 else
            "🟡 STRONG"   if s >= 60 else
            "🟠 GOOD"     if s >= 45 else
            "🔴 MODERATE" if s >= 30 else "⚫ WEAK"
        )
        tier_vals = out_df[under_label].apply(
            lambda x: u_tier(float(x)) if pd.notna(x) else "—"
        )
        insert_at = out_df.columns.get_loc(under_label) + 1
        out_df    = out_df.copy()          # decouple from disp before insert
        out_df.insert(insert_at, 'Tier', tier_vals)

    _text_cols = {'Tier', 'Status', 'Rest', 'Market Edge', 'TB Line', 'TB Under',
                  'TB Over', 'HR Odds', 'P.Grd', 'P.Grd ↑', 'Batter', 'Team', 'Pitcher'}

    fmt = {}
    for cn in out_df.columns:
        if cn in _text_cols:
            continue
        if cn in ['K%','BB%','Hit%','1B%','XB%','HR%']:
            fmt[cn] = "{:.1f}%"
        elif cn in ['🎯 Hit','🔥 XB','💣 HR', under_label]:
            fmt[cn] = "{:.1f}"
        elif cn == 'vsPit':
            fmt[cn] = "{:.0f}"
        elif cn == 'Hist AVG':
            fmt[cn] = "{:.3f}"
        elif cn == 'Hist PA':
            fmt[cn] = "{:.0f}"
        elif cn in ['BvP AVG','BvP OPS']:
            fmt[cn] = "{:.3f}"
        elif cn == 'BvP AB':
            fmt[cn] = "{:.0f}"

    styled = out_df.style.format(fmt, na_rep="—")

    # Under_Score gradient
    if under_label in out_df.columns:
        try:
            styled = styled.background_gradient(
                subset=[under_label], cmap='RdYlGn', vmin=20, vmax=80
            )
        except Exception:
            pass

    # Offensive scores: reversed gradient (low = good for under)
    for col, cmap in [('🎯 Hit','RdYlGn_r'),('🔥 XB','RdYlGn_r'),('💣 HR','RdYlGn_r')]:
        if col in out_df.columns:
            try:
                styled = styled.background_gradient(subset=[col], cmap=cmap, vmin=0, vmax=100)
            except Exception:
                pass

    if 'K%' in out_df.columns:
        try:
            styled = styled.background_gradient(subset=['K%'], cmap='Greens', vmin=10, vmax=40)
        except Exception:
            pass

    if 'BB%' in out_df.columns:
        try:
            styled = styled.background_gradient(subset=['BB%'], cmap='Greens', vmin=5, vmax=20)
        except Exception:
            pass

    if 'P.Grd' in out_df.columns:
        styled = styled.map(style_grade_cell, subset=['P.Grd'])

    # Hist AVG: reversed gradient — lower AVG = better under candidate (greener)
    if 'Hist AVG' in out_df.columns:
        try:
            styled = styled.background_gradient(
                subset=['Hist AVG'], cmap='RdYlGn_r', vmin=0.100, vmax=0.400
            )
        except Exception:
            pass

    # BvP AVG: reversed — low BvP AVG vs this pitcher = good for under
    if 'BvP AVG' in out_df.columns:
        try:
            styled = styled.background_gradient(
                subset=['BvP AVG'], cmap='RdYlGn_r', vmin=0.100, vmax=0.450
            )
        except Exception:
            pass

    # BvP OPS: reversed — low OPS vs this pitcher = struggles here = under signal
    if 'BvP OPS' in out_df.columns:
        try:
            styled = styled.background_gradient(
                subset=['BvP OPS'], cmap='RdYlGn_r', vmin=0.300, vmax=1.100
            )
        except Exception:
            pass

    # ── Column config ─────────────────────────────────────────────────────────
    col_cfg: dict = {}
    try:
        import streamlit as _st
        CC = _st.column_config
        col_cfg['Batter'] = CC.TextColumn("Batter", width="medium")
        col_cfg['Team']   = CC.TextColumn("Team",   width="small")
        col_cfg['Status'] = CC.TextColumn("Status", width="small",
                                          help="✅ = clean candidate · ⚠️ = disqualified (offsetting high score)")
        col_cfg['Rest']   = CC.TextColumn("Rest", width="small",
                                          help="Pitcher days rest since last start. "
                                               "✅ 4-5+d = normal/well-rested. ⚠️ 3d = short rest. ❌ ≤2d = very short.")
        col_cfg['Tier']   = CC.TextColumn("Tier", width="small",
                                          help="Under Score tier: ELITE ≥75 · STRONG ≥60 · GOOD ≥45")
        if 'P.Grd ↑' in out_df.columns or 'P.Grd' in out_df.columns:
            pgrd_col = 'P.Grd ↑' if 'P.Grd ↑' in out_df.columns else 'P.Grd'
            col_cfg[pgrd_col] = CC.TextColumn(
                pgrd_col, width="small",
                help="Pitcher grade. A+/A = GOOD for unders (elite pitchers suppress hits). "
                     "D/C = BAD for unders (weak pitchers give up bases freely)."
            )
        if under_label in out_df.columns:
            col_cfg[under_label] = CC.NumberColumn(
                under_label, format="%.1f", min_value=0, max_value=100,
                help="Under Score 0–100. Higher = stronger under candidate."
            )
        if 'Market Edge' in out_df.columns:
            col_cfg['Market Edge'] = CC.TextColumn(
                "Market Edge", width="small",
                help="⚡ EDGE = model likes under more than market. ✅ CONFIRMED = both agree."
            )
        for odds_col in ['TB Under','TB Over','HR Odds']:
            if odds_col in out_df.columns:
                col_cfg[odds_col] = CC.TextColumn(odds_col, width="small")
        if 'TB Line' in out_df.columns:
            col_cfg['TB Line'] = CC.TextColumn("TB Line", width="small",
                                               help="Sportsbook line: 0.5 or 1.5 total bases")
        if 'Hist PA' in out_df.columns:
            col_cfg['Hist PA'] = CC.NumberColumn(
                "Hist PA", format="%d", min_value=0,
                help="Plate appearances vs this pitcher type (BallPark Pal). "
                     "≥5 PA = signal is meaningful. <5 = small sample."
            )
        if 'Hist H' in out_df.columns:
            col_cfg['Hist H'] = CC.NumberColumn(
                "Hist H", format="%d", min_value=0,
                help="Hits in those plate appearances vs this pitcher type."
            )
        if 'Hist AVG' in out_df.columns:
            col_cfg['Hist AVG'] = CC.NumberColumn(
                "Hist AVG", format="%.3f",
                help="Batting average vs this pitcher type. "
                     "Below .245 (league avg) = historically struggles here = under signal."
            )
    except Exception:
        col_cfg = {}

    st.dataframe(styled, use_container_width=True,
                 column_config=col_cfg or None, hide_index=False)

    # Legend
    disq_note = {
        'xb':   f"Disqualified if HR_Score>{cfg['under_xb_disq_hr']:.0f} or Hit_Score>{cfg['under_xb_disq_hit']:.0f}",
        'tb15': f"Disqualified if XB_Score>{cfg['under_xb_disq_hr']:.0f} or HR_Score>{cfg['under_xb_disq_hr']:.0f} — singles are FINE for this line",
        'tb05': f"Disqualified if ANY of Hit/XB/HR_Score>{cfg['under_tb_disq_any']:.0f}",
        'hit':  f"Disqualified if Hit_Score>{cfg['under_hit_disq_hit']:.0f}",
        'hrr':  "Disqualified if Hit_Score>65 AND batting order slot 1-5 (cleanup hitters almost always accumulate H+R+RBI)",
    }.get(target, "")

    st.markdown(
        f'<div style="background:#0f1923;border:1px solid #1e2d3d;border-radius:8px;'
        f'padding:.6rem 1rem;font-size:.74rem;color:#64748b;margin-top:.4rem;line-height:1.7">'
        f'<b style="color:#94a3b8">Under Score</b> = weighted sum of (100 - offensive scores) '
        f'+ K% bonus + pitcher grade bonus. Higher = better under candidate.<br>'
        f'<b style="color:#94a3b8">⚠️ Disqualified:</b> {disq_note} — these players can still '
        f'accumulate bases through an offsetting route. Toggle "Show disqualified" to include.<br>'
        f'<b style="color:#94a3b8">Usage:</b> Cross-reference with sportsbook line. '
        f'Low offensive scores + strong pitcher + high K% = high-confidence under candidate.'
        f'</div>',
        unsafe_allow_html=True
    )


# ─────────────────────────────────────────────────────────────────────────────
# MAIN PAGE
# ─────────────────────────────────────────────────────────────────────────────

def under_page(df: pd.DataFrame, filters_base: dict):
    """
    Under Targets page.

    df          — already scored slate df (from app._build_scored_df)
    filters_base — base filters from the main sidebar (exclusions etc.
                   that persist across all pages via session_state)
    """
    st.title("🔻 Under Targets")
    st.markdown(
        '<div style="background:#0f1923;border:1px solid #3730a3;border-radius:8px;'
        'padding:.65rem 1rem;margin-bottom:.75rem;font-size:.82rem;color:#a5b4fc">'
        '📖 <b>How to use:</b> Find players unlikely to accumulate bases for a specific prop type. '
        'High Under Score = strong candidate. Always verify the sportsbook line before betting — '
        'an XB Under with score 78 means nothing if the book has it at -300 already. '
        'Disqualified (⚠️) players have a high score in an offsetting category that can still '
        'produce bases through a different route.</div>',
        unsafe_allow_html=True
    )

    if df is None or df.empty:
        st.error("❌ No slate data loaded.")
        return

    # Build sidebar filters FIRST — use_gc must be known before scoring
    filters = build_under_filters(df)

    # Compute under scores with form maps and GC from filters
    try:
        from mlb_api import get_recent_batting_form as _get_form, \
                            get_recent_pitcher_form as _get_pitcher_form, \
                            get_pitcher_rest_map    as _get_rest_map
        _form         = _get_form(days=7)
        _pitcher_form = _get_pitcher_form(days=7)
        _rest_map     = _get_rest_map()
    except Exception:
        _form         = {}
        _pitcher_form = {}
        _rest_map     = {}
    df = compute_under_scores(df, form_map=_form,
                              use_gc=filters.get('use_gc', True),
                              pitcher_form=_pitcher_form,
                              rest_map=_rest_map)

    # Apply confirmed lineup filter (same logic as main page)
    if filters.get('confirmed_only'):
        try:
            from mlb_api import get_confirmed_game_abbrs
            from config import NICK_TO_ABBR
            confirmed_abbrs = get_confirmed_game_abbrs()
            if confirmed_abbrs:
                def _game_confirmed(game_str: str) -> bool:
                    parts = str(game_str).split(' @ ')
                    if len(parts) != 2:
                        return False
                    away_abbr = NICK_TO_ABBR.get(parts[0].strip())
                    home_abbr = NICK_TO_ABBR.get(parts[1].strip())
                    return bool(away_abbr and home_abbr and
                                (away_abbr, home_abbr) in confirmed_abbrs)
                pre = len(df)
                df  = df[df['Game'].apply(_game_confirmed)]
                if len(df) < pre:
                    st.info(f"📋 {len(confirmed_abbrs)} confirmed games · "
                            f"{len(df)} batters shown")
        except Exception:
            pass

    if df.empty:
        st.warning("No players available after applying lineup filter.")
        return

    # Staleness warning
    if st.session_state.get('slate_stale'):
        slate_date = st.session_state.get('slate_date', 'unknown')
        st.markdown(
            f'<div style="background:#1c0000;border:2px solid #ef4444;border-radius:8px;'
            f'padding:.7rem 1rem;margin:.4rem 0;color:#f87171;font-size:.82rem">'
            f'⚠️ <b>Stale Data ({slate_date})</b> — today\'s slate not yet uploaded.</div>',
            unsafe_allow_html=True
        )

    # Full slate for top cards (only exclusions applied)
    slate_excl = df[~df['Batter'].isin(filters.get('excluded_players', []))]
    if filters.get('starters_only') and 'Starter' in slate_excl.columns:
        slate_excl = slate_excl[slate_excl['Starter'] == 1]
    if filters.get('include_teams'):
        slate_excl = slate_excl[slate_excl['Team'].isin(filters['include_teams'])]
    if filters.get('exclude_teams'):
        slate_excl = slate_excl[~slate_excl['Team'].isin(filters['exclude_teams'])]

    # Top cards
    st.markdown(
        '<div style="font-size:.68rem;text-transform:uppercase;letter-spacing:.1em;'
        'color:#5a7090;font-weight:700;margin:.8rem 0 .4rem;display:flex;align-items:center;gap:.5rem">'
        '🏆 BEST UNDER CANDIDATES — FULL SLATE'
        '<span style="flex:1;height:1px;background:linear-gradient(90deg,#1e2d3d,transparent);margin-left:.5rem"></span>'
        '</div>',
        unsafe_allow_html=True
    )
    _render_under_top_cards(slate_excl, filters)

    # ── Game Conditions Panel ─────────────────────────────────────────────────
    # Show today's game-level environment so users can see whether GC data
    # supports or conflicts with their under target.
    # Inverted from main predictor: LOW hits/runs/HR = ✅ for unders.
    # HIGH K% / QS% = ✅ for unders.
    _GC_DISPLAY = ['gc_hits20','gc_runs10','gc_k20','gc_qs','gc_hr4']
    if all(c in slate_excl.columns for c in _GC_DISPLAY):
        with st.expander("🌦️ Game Conditions — Under Context", expanded=False):
            st.markdown(
                '<div style="font-size:.74rem;color:#64748b;margin-bottom:.5rem">'
                '✅ = favors unders &nbsp;·&nbsp; ⚠️ = favors overs &nbsp;·&nbsp; '
                'Anchors: Hits 18.6% · Runs 28.4% · Ks 23.3% · QS 21.5% · HR 12.2%'
                '</div>',
                unsafe_allow_html=True
            )
            # Build one row per unique game
            game_rows = (slate_excl[['Game'] + _GC_DISPLAY]
                         .drop_duplicates(subset='Game')
                         .sort_values('Game'))
            cfg = CONFIG
            display_rows = []
            for _, row in game_rows.iterrows():
                def _flag_under(val, anchor, lower_good: bool) -> str:
                    """✅ when favorable for unders, ⚠️ when harmful."""
                    return "✅" if (val < anchor if lower_good else val > anchor) else "⚠️"
                display_rows.append({
                    'Game':       row['Game'],
                    '20+Hits %':  f"{_flag_under(row['gc_hits20'], cfg['gc_hits20_anchor'], True)}"
                                  f" {row['gc_hits20']:.1f}%",
                    '10+Runs %':  f"{_flag_under(row['gc_runs10'], cfg['gc_runs10_anchor'], True)}"
                                  f" {row['gc_runs10']:.1f}%",
                    '20+Ks %':    f"{_flag_under(row['gc_k20'],    cfg['gc_k20_anchor'],    False)}"
                                  f" {row['gc_k20']:.1f}%",
                    'SP QS %':    f"{_flag_under(row['gc_qs'],     cfg['gc_qs_anchor'],     False)}"
                                  f" {row['gc_qs']:.1f}%",
                    '4+HR %':     f"{_flag_under(row['gc_hr4'],    cfg['gc_hr4_anchor'],    True)}"
                                  f" {row['gc_hr4']:.1f}%",
                })
            if display_rows:
                st.dataframe(
                    pd.DataFrame(display_rows),
                    use_container_width=True, hide_index=True
                )

    # ── TB line auto-suggest from prop odds ────────────────────────────────────
    # If prop data is available, check whether the book's line matches the
    # user's selected under type — and surface a note if it doesn't.
    if 'prop_tb_line' in slate_excl.columns:
        lines_available = slate_excl['prop_tb_line'].dropna()
        lines_available = lines_available[lines_available.astype(bool)]
        if not lines_available.empty:
            line_counts = lines_available.value_counts()
            top_line    = line_counts.index[0]
            target      = filters['under']
            # Suggest switching if the dominant book line doesn't match selection
            if top_line == '0.5' and target == 'tb15':
                st.info(
                    "📊 **Book tip:** Most players today have TB lines at **0.5**. "
                    "Consider switching to **TB Under 0.5** for this slate."
                )
            elif top_line == '1.5' and target == 'tb05':
                st.info(
                    "📊 **Book tip:** Most players today have TB lines at **1.5**. "
                    "Consider switching to **TB Under 1.5** which matches the available market."
                )

    # Results table
    st.markdown("---")
    filtered_df = apply_under_filters(df, filters)

    target_lbl = filters['under_label'].split('—')[0].strip()
    st.markdown(
        f'<div style="display:flex;align-items:center;gap:.6rem;margin:.5rem 0 .35rem">'
        f'<span style="font-size:.9rem;font-weight:700;color:#e2e8f0">'
        f'{target_lbl} Candidates</span>'
        f'<span style="background:#131c2e;border:1px solid #2a3f57;border-radius:20px;'
        f'padding:.15rem .6rem;font-family:\'JetBrains Mono\',monospace;font-size:.72rem;'
        f'color:#5a7090">{len(filtered_df)} results</span>'
        f'</div>',
        unsafe_allow_html=True
    )

    _render_under_table(filtered_df, filters)

    # Export
    st.markdown("---")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("🔄 Refresh Data", key="under_refresh"):
            st.cache_data.clear()
            st.rerun()
    with c2:
        if not filtered_df.empty:
            from datetime import datetime
            import io, openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            # Build clean under export df
            _drop = ['pitch_hit_mult','pitch_hr_mult','pitch_walk_pen',
                     'rc_norm','rc_contrib','vs_mod','vs_contrib',
                     'xb_boost','hist_bonus','Starter',
                     'p_1b_base','p_xb_base','p_hr_base','p_k_base','p_bb_base',
                     'Hit_Score_base','Single_Score_base','XB_Score_base','HR_Score_base']
            _int_cols = [c for c in filtered_df.columns if c.startswith('_')
                         or (c.startswith('Under_') and c != filters['under_score_col'])
                         or c.startswith('disq')]
            export_cols = [c for c in filtered_df.columns
                           if c not in _drop and c not in _int_cols]
            export_df = filtered_df[export_cols].copy()
            score_col = filters['under_score_col']
            lmap = {
                'Under_XB_Score':'XB Under Score','Under_TB15_Score':'TB 1.5 Under Score',
                'Under_TB05_Score':'TB 0.5 Under Score','Under_Hit_Score':'Hit Under Score',
                'Under_HRR_Score':'H+R+RBI Under Score',
            }
            export_df = export_df.rename(columns={
                score_col: lmap.get(score_col,'Under Score'),
                'pitch_grade':'Pitcher Grade','p_k':'K%','p_bb':'BB%',
                'total_hit_prob':'Hit%','p_1b':'1B%','p_xb':'XB%','p_hr':'HR%',
                'vs Grade':'vs Pitcher Grade','Hit_Score':'Hit Score',
                'XB_Score':'XB Score','HR_Score':'HR Score','PA':'Hist PA',
                'H':'Hist H','AVG':'Hist AVG','bvp_avg':'BvP AVG',
                'bvp_ops':'BvP OPS','bvp_ab':'BvP AB',
                'prop_tb_line':'TB Line','prop_tb_under_odds':'TB Under Odds',
                'split_avg':'Split AVG',
            })
            ul = lmap.get(score_col, 'Under Score')

            # Round numerics
            for col in export_df.columns:
                try:
                    s = pd.to_numeric(export_df[col], errors='coerce')
                    if s.notna().any():
                        export_df[col] = s.round(3 if 'AVG' in col or 'OPS' in col else 1)
                except Exception:
                    pass

            try:
                # Build xlsx
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Under Targets"
                ws.sheet_properties.tabColor = "F87171"

                HDR_BG = "0D1321"; HDR_FG = "E8EEF7"
                ALT    = "0F1923"
                TIER = {"ELITE":("052E16","4ADE80"),"STRONG":("1C2A00","A3E635"),
                        "GOOD":("1C1500","FBBF24"),"MODERATE":("1C0800","FB923C"),
                        "WEAK":("1C0000","F87171")}

                hfill = PatternFill("solid", fgColor=HDR_BG)
                hfont = Font(bold=True, color=HDR_FG, name="Calibri", size=9)
                halign= Alignment(horizontal="center", vertical="center", wrap_text=True)

                for ci, cn in enumerate(export_df.columns, 1):
                    c = ws.cell(1, ci, cn)
                    c.fill=hfill; c.font=hfont; c.alignment=halign
                ws.row_dimensions[1].height = 28
                ws.freeze_panes = "A2"

                # Insert Tier column after under score
                tier_col_idx = None
                if ul in export_df.columns:
                    tier_col_idx = list(export_df.columns).index(ul) + 2  # 1-based after score

                bfont = Font(name="Calibri", size=9, color=HDR_FG)
                afill = PatternFill("solid", fgColor=ALT)

                for ri, (_, row) in enumerate(export_df.iterrows(), 2):
                    for ci, (cn, val) in enumerate(row.items(), 1):
                        cell = ws.cell(ri, ci, val)
                        cell.font = bfont
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        if ri % 2 == 0:
                            cell.fill = afill
                        if cn == "Tier" and val in TIER:
                            bg, fg = TIER[val]
                            cell.fill = PatternFill("solid", fgColor=bg)
                            cell.font = Font(name="Calibri", size=9, color=fg, bold=True)
                        elif "AVG" in cn or "OPS" in cn:
                            cell.number_format = "0.000"
                        elif "Score" in cn or "%" in cn:
                            cell.number_format = "0.0"

                ws.auto_filter.ref = f"A1:{get_column_letter(len(export_df.columns))}1"

                # Col widths
                for ci, cn in enumerate(export_df.columns, 1):
                    l = get_column_letter(ci)
                    ws.column_dimensions[l].width = (
                        16 if cn in ("Batter","Pitcher") else
                        13 if "Score" in cn else
                        10 if "AVG" in cn or "OPS" in cn or "BvP" in cn else
                        8  if "%" in cn else 9
                    )

                buf = io.BytesIO(); wb.save(buf); buf.seek(0)
                st.download_button(
                    "📊 Export Excel",
                    buf.read(),
                    f"a1picks_unders_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="under_export",
                    help="Color-coded Excel with tier highlights and BvP formatting",
                )
            except Exception:
                st.download_button(
                    "💾 Export CSV (fallback)",
                    export_df.to_csv(index=False),
                    f"a1picks_unders_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                    mime="text/csv", key="under_export_csv",
                )
